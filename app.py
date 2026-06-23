print(">>> APP.PY CORRECTO CARGADO <<<")
import os
from werkzeug.utils import secure_filename
from flask import send_file, send_from_directory, jsonify
from openpyxl import Workbook, load_workbook
from io import BytesIO
from flask import Flask, render_template, request, redirect, url_for, session, send_file
from cajas import (
    crear_caja,
    eliminar_caja,
    modificar_caja,
    asegurar_caja_sin_asignar,
    reparar_archivos_huerfanos,
    reubicar_archivos_de_caja,
    reubicar_archivos_pendientes_por_nueva_caja
)
from archivos import agregar_archivo, modificar_archivo, eliminar_archivo
from auth import crear_usuario, verificar_usuario, obtener_usuario_y_rol, usuario_existe
from grupos import (
    crear_grupo,
    agregar_usuario_a_grupo,
    obtener_grupos_usuario,
    obtener_todos_grupos,
    obtener_miembros_grupo,
    buscar_usuario_por_nombre,
    usuario_puede_eliminar,
    quitar_usuario_de_grupo,
    sincronizar_registros_usuario,
    eliminar_grupo_personal,
    archivar_grupo
)
from logs import registrar_log
from historial import registrar_movimiento
from werkzeug.security import generate_password_hash
from db import get_db
from io import BytesIO
from datetime import datetime
from zoneinfo import ZoneInfo
import json
from collections import defaultdict
from difflib import SequenceMatcher
from openpyxl.styles import Font
from flask import flash
import threading
import uuid
import re
import zipfile
from werkzeug.exceptions import HTTPException
try:
    from PyPDF2 import PdfReader, PdfWriter
except Exception:
    PdfReader = None
    PdfWriter = None

from error_catalog import error_text

app = Flask(__name__)
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
UPLOAD_FOLDER = os.path.join(BASE_DIR, "uploads", "pdfs")
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER
app.config["MAX_CONTENT_LENGTH"] = 260 * 1024 * 1024  # 260MB total request size
app.config["MASSIVE_PDF_MAX_FILES"] = 500
app.config["MASSIVE_PDF_MAX_TOTAL_BYTES"] = 250 * 1024 * 1024

app.secret_key = os.environ.get("FLASK_SECRET_KEY", "clave_super_secreta")
app.config["SESSION_COOKIE_NAME"] = os.environ.get("FLASK_SESSION_COOKIE_NAME", "session")
app.config["SESSION_COOKIE_HTTPONLY"] = True
app.config["SESSION_COOKIE_SAMESITE"] = "Lax"
app.config["SESSION_COOKIE_SECURE"] = os.environ.get("FLASK_SECURE_COOKIES") == "1"
app.config["ADMIN_BULK_DELETE_PASSWORD"] = os.environ.get("ADMIN_BULK_DELETE_PASSWORD", "1122514048")
APP_TIMEZONE = ZoneInfo("America/Bogota")

IMPORT_JOBS = {}
IMPORT_JOBS_LOCK = threading.Lock()
IMPORT_REPORTS_TABLE_READY = False


def flash_error(code, fallback=None, detail=None):
    flash(error_text(code, fallback=fallback, detail=detail), "error")


def ensure_import_reports_table():
    global IMPORT_REPORTS_TABLE_READY
    if IMPORT_REPORTS_TABLE_READY:
        return
    conn = get_db()
    cur = conn.cursor()
    cur.execute(
        """
        IF OBJECT_ID('dbo.importaciones_excel', 'U') IS NULL
        BEGIN
            CREATE TABLE dbo.importaciones_excel (
                id INT IDENTITY(1,1) PRIMARY KEY,
                job_id NVARCHAR(64) NOT NULL UNIQUE,
                import_type NVARCHAR(20) NOT NULL CONSTRAINT DF_importaciones_excel_import_type DEFAULT ('excel'),
                usuarioid INT NULL,
                nombreusuario NVARCHAR(255) NULL,
                empresa INT NULL,
                nombreempresa NVARCHAR(255) NULL,
                archivo_nombre NVARCHAR(255) NULL,
                status NVARCHAR(20) NOT NULL,
                error_code INT NULL,
                total_rows INT NOT NULL CONSTRAINT DF_importaciones_excel_total_rows DEFAULT (0),
                processed_rows INT NOT NULL CONSTRAINT DF_importaciones_excel_processed_rows DEFAULT (0),
                inserted INT NOT NULL CONSTRAINT DF_importaciones_excel_inserted DEFAULT (0),
                merged INT NOT NULL CONSTRAINT DF_importaciones_excel_merged DEFAULT (0),
                ignored INT NOT NULL CONSTRAINT DF_importaciones_excel_ignored DEFAULT (0),
                invalid INT NOT NULL CONSTRAINT DF_importaciones_excel_invalid DEFAULT (0),
                invalid_details NVARCHAR(MAX) NULL,
                detail NVARCHAR(MAX) NULL,
                started_at DATETIME2 NULL,
                finished_at DATETIME2 NULL,
                created_at DATETIME2 NOT NULL CONSTRAINT DF_importaciones_excel_created_at DEFAULT (SYSDATETIME())
            );
        END

        IF COL_LENGTH('dbo.importaciones_excel', 'import_type') IS NULL
        BEGIN
            ALTER TABLE dbo.importaciones_excel
            ADD import_type NVARCHAR(20) NOT NULL
                CONSTRAINT DF_importaciones_excel_import_type_legacy DEFAULT ('excel');
        END

        IF COL_LENGTH('dbo.importaciones_excel', 'merged') IS NULL
        BEGIN
            ALTER TABLE dbo.importaciones_excel
            ADD merged INT NOT NULL
                CONSTRAINT DF_importaciones_excel_merged_legacy DEFAULT (0);
        END
        """
    )
    conn.commit()
    cur.close()
    conn.close()
    IMPORT_REPORTS_TABLE_READY = True


def _json_loads_safe(value, default):
    if not value:
        return default
    try:
        return json.loads(value)
    except Exception:
        return default


def _save_import_report(job):
    if not job or not job.get("job_id"):
        return
    ensure_import_reports_table()
    conn = get_db()
    cur = conn.cursor()
    invalid_details_text = json.dumps(job.get("invalid_details") or [], ensure_ascii=False)
    cur.execute("SELECT 1 FROM importaciones_excel WHERE job_id = %s", (job["job_id"],))
    exists = cur.fetchone()
    if exists:
        cur.execute(
            """
            UPDATE importaciones_excel
            SET import_type = %s,
                usuarioid = %s,
                nombreusuario = %s,
                empresa = %s,
                nombreempresa = %s,
                archivo_nombre = %s,
                status = %s,
                error_code = %s,
                total_rows = %s,
                processed_rows = %s,
                inserted = %s,
                merged = %s,
                ignored = %s,
                invalid = %s,
                invalid_details = %s,
                detail = %s,
                started_at = %s,
                finished_at = %s
            WHERE job_id = %s
            """,
            (
                job.get("import_type", "excel"),
                job.get("user_id"),
                job.get("user_name"),
                job.get("group_id"),
                job.get("group_name"),
                job.get("source_filename"),
                job.get("status"),
                job.get("error_code"),
                job.get("total_rows", 0),
                job.get("processed_rows", 0),
                job.get("inserted", 0),
                job.get("merged", 0),
                job.get("ignored", 0),
                job.get("invalid", 0),
                invalid_details_text,
                job.get("detail", ""),
                job.get("started_at"),
                job.get("finished_at"),
                job["job_id"],
            ),
        )
    else:
        cur.execute(
            """
            INSERT INTO importaciones_excel (
                job_id, import_type, usuarioid, nombreusuario, empresa, nombreempresa, archivo_nombre,
                status, error_code, total_rows, processed_rows, inserted, merged, ignored, invalid,
                invalid_details, detail, started_at, finished_at
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """,
            (
                job["job_id"],
                job.get("import_type", "excel"),
                job.get("user_id"),
                job.get("user_name"),
                job.get("group_id"),
                job.get("group_name"),
                job.get("source_filename"),
                job.get("status"),
                job.get("error_code"),
                job.get("total_rows", 0),
                job.get("processed_rows", 0),
                job.get("inserted", 0),
                job.get("merged", 0),
                job.get("ignored", 0),
                job.get("invalid", 0),
                invalid_details_text,
                job.get("detail", ""),
                job.get("started_at"),
                job.get("finished_at"),
            ),
        )
    conn.commit()
    cur.close()
    conn.close()


def persist_import_report(job):
    _save_import_report(job)


def get_import_report(job_id):
    ensure_import_reports_table()
    conn = get_db()
    cur = conn.cursor()
    cur.execute(
        """
        SELECT
            job_id, import_type, usuarioid, nombreusuario, empresa, nombreempresa, archivo_nombre,
            status, error_code, total_rows, processed_rows, inserted, merged, ignored, invalid,
            invalid_details, detail, started_at, finished_at, created_at
        FROM importaciones_excel
        WHERE job_id = %s
        """,
        (job_id,)
    )
    row = cur.fetchone()
    cur.close()
    conn.close()
    if not row:
        return None
    return {
        "job_id": row[0],
        "import_type": row[1] or "excel",
        "user_id": row[2],
        "user_name": row[3],
        "group_id": row[4],
        "group_name": row[5],
        "source_filename": row[6],
        "status": row[7],
        "error_code": row[8],
        "total_rows": row[9],
        "processed_rows": row[10],
        "inserted": row[11],
        "merged": row[12],
        "ignored": row[13],
        "invalid": row[14],
        "invalid_details": _json_loads_safe(row[15], []),
        "detail": row[16] or "",
        "started_at": row[17].isoformat() if row[17] else None,
        "finished_at": row[18].isoformat() if row[18] else None,
        "created_at": row[19].isoformat() if row[19] else None,
    }


def get_recent_import_reports(limit=10, group_id=None):
    ensure_import_reports_table()
    conn = get_db()
    cur = conn.cursor()
    limit = max(1, int(limit))
    if group_id:
        cur.execute(
            f"""
            SELECT TOP {limit}
                job_id, import_type, nombreusuario, nombreempresa, archivo_nombre, status, error_code,
                total_rows, processed_rows, inserted, merged, ignored, invalid, started_at, finished_at
            FROM importaciones_excel
            WHERE empresa = %s
            ORDER BY COALESCE(finished_at, started_at, created_at) DESC, id DESC
            """,
            (group_id,)
        )
    else:
        cur.execute(
            f"""
            SELECT TOP {limit}
                job_id, import_type, nombreusuario, nombreempresa, archivo_nombre, status, error_code,
                total_rows, processed_rows, inserted, merged, ignored, invalid, started_at, finished_at
            FROM importaciones_excel
            ORDER BY COALESCE(finished_at, started_at, created_at) DESC, id DESC
            """
        )
    rows = cur.fetchall()
    cur.close()
    conn.close()
    return [
        {
            "job_id": r[0],
            "import_type": r[1] or "excel",
            "user_name": r[2],
            "group_name": r[3],
            "source_filename": r[4],
            "status": r[5],
            "error_code": r[6],
            "total_rows": r[7],
            "processed_rows": r[8],
            "inserted": r[9],
            "merged": r[10],
            "ignored": r[11],
            "invalid": r[12],
            "started_at": r[13].isoformat() if r[13] else None,
            "finished_at": r[14].isoformat() if r[14] else None,
        }
        for r in rows
    ]


def get_group_name(group_id):
    if not group_id:
        return ""
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT nombre FROM empresas WHERE id = %s", (group_id,))
    row = cur.fetchone()
    cur.close()
    conn.close()
    return row[0] if row else ""


def puede_ver_reporte_importacion(job):
    if not job:
        return False
    if admin_requerido():
        return True
    grupo_actual = obtener_grupo_id()
    return bool(grupo_actual and job.get("group_id") == grupo_actual)


def build_import_job_payload(job_id, **extra):
    payload = {
        "job_id": job_id,
        "import_type": "excel",
        "status": "pending",
        "message": "Importacion en cola.",
        "error_code": None,
        "total_rows": 0,
        "processed_rows": 0,
        "inserted": 0,
        "merged": 0,
        "ignored": 0,
        "invalid": 0,
        "invalid_details": [],
        "detail": "",
        "source_filename": "",
        "user_id": None,
        "user_name": "",
        "group_id": None,
        "group_name": "",
        "started_at": datetime.utcnow().isoformat(),
        "created_at": datetime.utcnow().isoformat(),
        "finished_at": None,
        "report_saved": False,
    }
    payload.update(extra)
    return payload


def get_import_labels(import_type):
    if str(import_type or "excel").lower() == "pdf":
        return {
            "kind": "pdf",
            "item": "PDF",
            "plural": "PDFs",
            "success_label": "nuevos",
            "merged_label": "unidos",
            "ignored_label": "no encontrados",
        }
    return {
        "kind": "excel",
        "item": "Excel",
        "plural": "filas",
        "success_label": "exitosos",
        "merged_label": "unidos",
        "ignored_label": "repetidos",
    }


def set_import_job(job_id, **updates):
    with IMPORT_JOBS_LOCK:
        current = IMPORT_JOBS.get(job_id, build_import_job_payload(job_id))
        current.update(updates)
        IMPORT_JOBS[job_id] = current
        snapshot = dict(current)
    try:
        _save_import_report(snapshot)
    except Exception:
        app.logger.exception("No se pudo persistir el estado de importacion %s", job_id)
    return snapshot


def get_import_job(job_id):
    with IMPORT_JOBS_LOCK:
        job = IMPORT_JOBS.get(job_id)
        return dict(job) if job else None


def add_invalid_detail(invalid_details, value, row_number=None, reason=None, max_items=150):
    if len(invalid_details) >= max_items:
        return
    raw = "" if value is None else str(value).strip()
    if not raw:
        raw = f"fila {row_number}" if row_number is not None else "valor vacio"
    detail = raw
    if row_number is not None:
        detail = f"Fila {row_number}: {detail}"
    if reason:
        detail = f"{detail} ({reason})"
    invalid_details.append(detail)

def csrf_token():
    token = session.get("_csrf_token")
    if not token:
        token = os.urandom(16).hex()
        session["_csrf_token"] = token
    return token

@app.context_processor
def inject_csrf_token():
    return {"csrf_token": csrf_token}

@app.before_request
def csrf_protect():
    if request.method == "POST":
        sent = request.form.get("_csrf_token")
        if not sent or sent != session.get("_csrf_token"):
            flash_error(1)
            return redirect(request.referrer or url_for("inicio"))

app.jinja_env.globals["csrf_token"] = csrf_token


@app.route("/importacion/estado/<job_id>")
def importacion_estado(job_id):
    if not login_requerido():
        return jsonify({"ok": False, "error": "Debes iniciar sesion para continuar."}), 401

    job = get_import_job(job_id)
    if not job:
        job = get_import_report(job_id)
    if not job:
        return jsonify({"ok": False, "error": error_text(750)}), 404

    if not puede_ver_reporte_importacion(job):
        return jsonify({"ok": False, "error": error_text(206)}), 403

    return jsonify({"ok": True, "job": job})


@app.route("/importacion/reporte/<job_id>")
def importacion_reporte(job_id):
    if not login_requerido():
        return jsonify({"ok": False, "error": "Debes iniciar sesion para continuar."}), 401

    job = get_import_job(job_id) or get_import_report(job_id)
    if not job:
        return jsonify({"ok": False, "error": error_text(750)}), 404

    if not puede_ver_reporte_importacion(job):
        return jsonify({"ok": False, "error": error_text(206)}), 403

    return jsonify({"ok": True, "report": job})


@app.route("/importacion/cerrar", methods=["POST"])
def importacion_cerrar():
    if not login_requerido():
        return jsonify({"ok": False, "error": "Debes iniciar sesion para continuar."}), 401

    job_id = (request.form.get("job_id") or "").strip()
    current_job = session.get("last_import_job_id")

    if not job_id or current_job == job_id:
        session.pop("last_import_job_id", None)

    return jsonify({"ok": True})


@app.route("/informes")
def informes():
    if not login_requerido():
        return redirect(url_for("login"))

    grupo_id = obtener_grupo_id()
    if not grupo_id:
        return redirect(url_for("grupos"))

    selected_group_id = grupo_id
    group_options = []

    if admin_requerido():
        requested_group = request.args.get("empresa", "").strip()
        if requested_group and requested_group.isdigit():
            selected_group_id = int(requested_group)
        group_options = obtener_todos_grupos()

    reports = get_recent_import_reports(100, group_id=selected_group_id)
    selected_group_name = get_group_name(selected_group_id)

    return render_template(
        "informes.html",
        reports=reports,
        selected_group_id=selected_group_id,
        selected_group_name=selected_group_name,
        group_options=group_options,
        import_report_url_base=url_for("importacion_reporte", job_id="__JOB__"),
    )

def normalizar_nombre(nombre):
    return nombre.upper() if nombre else nombre

def normalizar_nombre_clave(nombre):
    if not nombre:
        return ""
    limpio = []
    for ch in str(nombre).upper().strip():
        if ch.isalnum() or ch.isspace():
            limpio.append(ch)
    base = "".join(limpio)
    return " ".join(base.split())


def json_text(value):
    return json.dumps(value, ensure_ascii=False) if value is not None else None


def json_or_none(value):
    if value is None:
        return None
    if isinstance(value, (dict, list)):
        return value
    try:
        return json.loads(value)
    except Exception:
        return None


def rows_to_builtin(rows):
    return [list(row) for row in rows]


def insert_with_identity(cur, table_name, sql, params):
    cur.execute(f"SET IDENTITY_INSERT {table_name} ON")
    try:
        cur.execute(sql, params)
    finally:
        cur.execute(f"SET IDENTITY_INSERT {table_name} OFF")


def sqlserver_in_clause(values):
    values = list(values)
    if not values:
        return "(NULL)", tuple()
    return "(" + ", ".join(["%s"] * len(values)) + ")", tuple(values)


def iter_chunks(values, size=900):
    values = list(values)
    for i in range(0, len(values), size):
        yield values[i:i + size]


def parse_item_movimiento(item):
    if not item:
        return None, None
    if ":" not in item:
        return item, None
    entidad, entidad_id = item.split(":", 1)
    try:
        return entidad, int(entidad_id)
    except ValueError:
        return entidad, None


def insertar_archivos_lote(cur, rows):
    insertados = []
    ignorados = 0
    for caja_id, numero, nombre, grupo_id, creado_por, tipo_doc in rows:
        cur.execute(
            "SELECT TOP 1 id FROM archivos WHERE grupo_id = %s AND numero = %s",
            (grupo_id, numero)
        )
        if cur.fetchone():
            ignorados += 1
            continue
        cur.execute(
            """
            INSERT INTO archivos (caja_id, numero, nombre, grupo_id, creado_por, tipo_doc)
            VALUES (%s, %s, %s, %s, %s, %s)
            RETURNING id
            """,
            (caja_id, numero, nombre, grupo_id, creado_por, tipo_doc)
        )
        archivo_id = cur.fetchone()[0]
        insertados.append((archivo_id, numero, nombre, caja_id, tipo_doc))
    return insertados, ignorados


def insertar_movimientos_lote(cur, rows):
    for usuario_id, grupo_id, entidad, entidad_id, accion, datos_antes, datos_despues, meta in rows:
        item = f"{entidad}:{entidad_id}" if entidad_id is not None else entidad
        antes_texto = datos_antes if isinstance(datos_antes, str) else json_text(datos_antes) or ""
        despues_base = datos_despues if isinstance(datos_despues, str) else json_text(datos_despues)
        meta_texto = meta if isinstance(meta, str) else json_text(meta)
        despues_texto = despues_base or meta_texto or ""
        cur.execute(
            """
            INSERT INTO movimientos (
                usuarioid, empresa, accion, antes, despues, item
            )
            VALUES (%s, %s, %s, %s, %s, %s)
            """,
            (usuario_id, grupo_id, accion, antes_texto, despues_texto, item)
        )

TIPO_DOC_OPCIONES = ("CC", "CE", "TI", "RC")

def normalizar_tipo_doc(tipo_doc):
    return tipo_doc.upper().strip() if tipo_doc else ""

def es_tipo_doc_valido(tipo_doc):
    return tipo_doc in TIPO_DOC_OPCIONES

@app.template_filter("puntos")
def formato_puntos(valor):
    try:
        numero = int(valor)
    except (TypeError, ValueError):
        return valor
    return f"{numero:,}".replace(",", ".")


@app.template_filter("hora_col")
def formato_hora_colombia(valor):
    if not valor:
        return "N/D"
    dt = valor
    if isinstance(valor, str):
        texto = valor.strip()
        if not texto:
            return "N/D"
        try:
            dt = datetime.fromisoformat(texto.replace("Z", "+00:00"))
        except Exception:
            return valor
    if not isinstance(dt, datetime):
        return valor
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=ZoneInfo("UTC"))
    return dt.astimezone(APP_TIMEZONE).strftime("%d/%m/%Y, %I:%M:%S %p")

def es_pdf(file):
    if not file:
        return False
    filename = (file.filename or "").lower()
    return filename.endswith(".pdf")


def obtener_tamano_upload(file):
    if not file or not getattr(file, "stream", None):
        return 0
    try:
        actual = file.stream.tell()
    except Exception:
        actual = 0
    try:
        file.stream.seek(0, os.SEEK_END)
        total = file.stream.tell()
        file.stream.seek(actual)
        return max(0, int(total))
    except Exception:
        try:
            file.stream.seek(0)
        except Exception:
            pass
        return 0

def guardar_pdf(file, numero_documento, grupo_id):
    """
    Guarda el pdf y retorna el path relativo (para guardar en DB).
    """
    filename = secure_filename(file.filename)
    # Guardar con nombre controlado (evita duplicados raros)
    final_name = f"doc_{grupo_id}_{numero_documento}.pdf"
    abs_path = os.path.join(app.config["UPLOAD_FOLDER"], final_name)
    file.save(abs_path)
    # Guardamos en DB un path relativo simple
    return final_name


def unir_pdf_existente(pdf_old, file, numero_documento, grupo_id):
    """
    Une el PDF existente con uno nuevo y guarda en el mismo archivo.
    Retorna el path relativo final.
    """
    if not pdf_old:
        return guardar_pdf(file, numero_documento, grupo_id)
    if PdfReader is None or PdfWriter is None:
        return None

    path = pdf_old
    if not os.path.isabs(path):
        path = os.path.join(app.config["UPLOAD_FOLDER"], path)

    if not os.path.exists(path):
        return guardar_pdf(file, numero_documento, grupo_id)

    try:
        reader_old = PdfReader(path)
        file.stream.seek(0)
        reader_new = PdfReader(file.stream)
        writer = PdfWriter()
        for p in reader_old.pages:
            writer.add_page(p)
        for p in reader_new.pages:
            writer.add_page(p)
        with open(path, "wb") as f:
            writer.write(f)
        return pdf_old
    except Exception:
        return None


def obtener_numero_desde_nombre_pdf(filename):
    if not filename:
        return None
    base = os.path.basename(filename)
    nombre, ext = os.path.splitext(base)
    if ext.lower() != ".pdf":
        return None
    nombre = nombre.strip()
    prefixed = re.fullmatch(r"\d{4}_(\d+)", nombre)
    if prefixed:
        nombre = prefixed.group(1)
    if not re.fullmatch(r"\d+", nombre):
        return None
    try:
        return int(nombre)
    except Exception:
        return None


def obtener_nombre_original_pdf(filename):
    if not filename:
        return ""
    base = os.path.basename(filename)
    nombre, ext = os.path.splitext(base)
    if ext.lower() != ".pdf":
        return base
    nombre = nombre.strip()
    prefixed = re.fullmatch(r"\d{4}_(.+)", nombre)
    if prefixed:
        return f"{prefixed.group(1)}{ext.lower()}"
    return base


def login_requerido():
    return "usuario" in session


def admin_requerido():
    return session.get("rol") == "admin"


def obtener_grupo_id():
    return session.get("grupo_id")


def supervisor_requerido():
    return session.get("rol") in ("admin", "supervisor")


def puede_crear_modificar_cajas():
    return session.get("rol") in ("admin", "supervisor", "usuario")


def puede_eliminar_cajas():
    return session.get("rol") in ("admin", "supervisor")


def usuario_requerido():
    return session.get("rol") not in ("admin", "supervisor")


def es_archivador_grupo(grupo_id):
    if not grupo_id:
        return False
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT nombre FROM grupos WHERE id = %s", (grupo_id,))
    row = cur.fetchone()
    cur.close()
    conn.close()
    return row and row[0] == "Archivador"


def importar_excel_job(file_path, grupo_id, usuario_id, job_id):
    """
    Procesa la importacion de Excel en segundo plano con inserciones por lotes.
    """
    set_import_job(
        job_id,
        status="processing",
        message="Importacion en proceso.",
        processed_rows=0,
        inserted=0,
        ignored=0,
        invalid=0,
        invalid_details=[],
        error_code=None,
        finished_at=None,
        report_saved=False,
    )
    try:
        wb = load_workbook(file_path, data_only=True)
    except Exception:
        app.logger.exception("Archivo Excel invalido")
        set_import_job(
            job_id,
            status="failed",
            message=error_text(501),
            error_code=501,
            finished_at=datetime.utcnow().isoformat(),
            report_saved=True,
        )
        persist_import_report(get_import_job(job_id))
        try:
            os.remove(file_path)
        except Exception:
            pass
        return

    ws = wb.active
    max_row = ws.max_row or 1
    max_col = ws.max_column or 1
    total_rows = max(max_row - 1, 0)
    set_import_job(job_id, total_rows=total_rows)

    headers = []
    for col in range(1, max_col + 1):
        val = ws.cell(row=1, column=col).value
        headers.append(str(val).strip().lower() if val is not None else "")

    def _clean_num(value):
        if value is None:
            return None
        if isinstance(value, (int, float)):
            try:
                return int(value)
            except Exception:
                return None
        s = str(value).strip()
        if not s:
            return None
        s = s.replace(".", "")
        if not s.isdigit():
            return None
        return int(s)

    def _clean_text(value):
        if value is None:
            return ""
        return str(value).strip()

    def _build_nombre(parts):
        return " ".join([p for p in parts if p])

    # ===== Formato nuevo (filas con tipo_doc + documento + nombres/apellidos)
    # Por decision operativa, toda importacion entra primero a Caja 0.
    if "tipo_doc" in headers and "documento" in headers:
        tipo_idx = headers.index("tipo_doc") + 1
        doc_idx = headers.index("documento") + 1
        pri_nom_idx = headers.index("pri_nom") + 1 if "pri_nom" in headers else None
        seg_nom_idx = headers.index("seg_nom") + 1 if "seg_nom" in headers else None
        pri_ap_idx = headers.index("pri_apelli") + 1 if "pri_apelli" in headers else None
        seg_ap_idx = headers.index("seg_apelli") + 1 if "seg_apelli" in headers else None

        conn = get_db()
        cur = conn.cursor()
        try:
            caja_pendiente_id = asegurar_caja_sin_asignar(grupo_id)

            values = []
            invalidos = 0
            invalid_details = []
            for row in range(2, max_row + 1):
                tipo_doc = normalizar_tipo_doc(ws.cell(row=row, column=tipo_idx).value)
                numero_raw = ws.cell(row=row, column=doc_idx).value
                if not es_tipo_doc_valido(tipo_doc):
                    invalidos += 1
                    add_invalid_detail(invalid_details, numero_raw, row_number=row, reason="tipo_doc invalido")
                    continue
                numero = _clean_num(numero_raw)
                if not numero:
                    invalidos += 1
                    add_invalid_detail(invalid_details, numero_raw, row_number=row, reason="documento invalido")
                    continue

                parts = []
                if pri_nom_idx:
                    parts.append(_clean_text(ws.cell(row=row, column=pri_nom_idx).value))
                if seg_nom_idx:
                    parts.append(_clean_text(ws.cell(row=row, column=seg_nom_idx).value))
                if pri_ap_idx:
                    parts.append(_clean_text(ws.cell(row=row, column=pri_ap_idx).value))
                if seg_ap_idx:
                    parts.append(_clean_text(ws.cell(row=row, column=seg_ap_idx).value))

                nombre = _build_nombre(parts)
                if not nombre:
                    nombre = f"Documento {numero}"
                nombre = normalizar_nombre(nombre)

                values.append((caja_pendiente_id, numero, nombre, grupo_id, usuario_id, tipo_doc))

            if not values:
                app.logger.warning("Importacion Excel: sin filas validas")
                set_import_job(
                    job_id,
                    status="partial",
                    message=error_text(503, detail="No se encontraron filas validas."),
                    processed_rows=total_rows,
                    inserted=0,
                    ignored=0,
                    invalid=invalidos,
                    invalid_details=invalid_details,
                    error_code=503,
                    finished_at=datetime.utcnow().isoformat(),
                    report_saved=True,
                )
                persist_import_report(get_import_job(job_id))
                return

            inserted = 0
            ignorados = 0
            procesados = invalidos
            for chunk in iter_chunks(values, size=500):
                rows, ignored_chunk = insertar_archivos_lote(cur, chunk)
                ignorados += ignored_chunk
                inserted += len(rows)
                procesados += len(chunk)

                if rows:
                    mov_values = [
                        (
                            usuario_id,
                            grupo_id,
                            "archivo",
                            r[0],
                            "CREAR_ARCHIVO",
                            None,
                            json_text({"id": r[0], "numero": r[1], "nombre": r[2], "caja_id": r[3], "tipo_doc": r[4]}),
                            None,
                        )
                        for r in rows
                    ]
                    insertar_movimientos_lote(cur, mov_values)

                conn.commit()
                set_import_job(
                    job_id,
                    processed_rows=min(procesados, total_rows),
                    inserted=inserted,
                    ignored=ignorados,
                    invalid=invalidos,
                    invalid_details=invalid_details,
                    detail=f"Nuevos archivos: {inserted} | Ignorados: {ignorados} | Invalidos: {invalidos}",
                )

            final_status = "success" if not invalidos and not ignorados else "partial"
            detail = f"Nuevos archivos: {inserted} | Ignorados: {ignorados} | Invalidos: {invalidos}"
            set_import_job(
                job_id,
                status=final_status,
                message=(
                    "Importacion finalizada correctamente."
                    if final_status == "success"
                    else "Importacion finalizada con observaciones."
                ),
                processed_rows=total_rows,
                inserted=inserted,
                ignored=ignorados,
                invalid=invalidos,
                invalid_details=invalid_details,
                error_code=None if final_status == "success" else 503,
                detail=detail,
                finished_at=datetime.utcnow().isoformat(),
                report_saved=True,
            )
            persist_import_report(get_import_job(job_id))
        except Exception:
            conn.rollback()
            app.logger.exception("Error importando Excel (formato filas)")
            set_import_job(
                job_id,
                status="failed",
                message=error_text(504),
                processed_rows=total_rows,
                error_code=504,
                finished_at=datetime.utcnow().isoformat(),
                report_saved=True,
            )
            persist_import_report(get_import_job(job_id))
        finally:
            cur.close()
            conn.close()
            try:
                os.remove(file_path)
            except Exception:
                pass
        return

    # ===== Formato anterior
    # Aunque el Excel venga separado por columnas, no se crean cajas desde la importacion.
    # Todo entra a Caja 0 y luego las cajas reales reubican los documentos automaticamente.
    columnas = []
    invalidos = 0
    invalid_details = []
    for col in range(1, max_col + 1):
        header = ws.cell(row=1, column=col).value
        if header is None or str(header).strip() == "":
            continue
        numeros = []
        for row in range(2, max_row + 1):
            cell_val = ws.cell(row=row, column=col).value
            num = _clean_num(cell_val)
            if not num:
                if cell_val not in (None, ""):
                    invalidos += 1
                    add_invalid_detail(invalid_details, cell_val, row_number=row, reason="documento invalido")
                continue
            numeros.append(num)
        columnas.append((col, str(header).strip(), numeros))

    if not columnas:
        app.logger.warning("Importacion Excel: sin cajas")
        set_import_job(
            job_id,
            status="partial",
            message=error_text(503, detail="El Excel no contenia columnas validas."),
            processed_rows=total_rows,
            inserted=0,
            ignored=0,
            invalid=invalidos or total_rows,
            invalid_details=invalid_details,
            error_code=503,
            finished_at=datetime.utcnow().isoformat(),
            report_saved=True,
        )
        persist_import_report(get_import_job(job_id))
        return

    conn = get_db()
    cur = conn.cursor()
    try:
        caja_pendiente_id = asegurar_caja_sin_asignar(grupo_id)
        numeros = []
        for _, _, nums in columnas:
            numeros.extend(nums)

        numeros = sorted(set(numeros))
        inserted = 0
        ignorados = 0
        procesados = invalidos
        for chunk in iter_chunks(numeros, size=500):
            rows, ignored_chunk = insertar_archivos_lote(
                cur,
                [(caja_pendiente_id, n, f"Documento {n}", grupo_id, usuario_id, "CC") for n in chunk]
            )
            ignorados += ignored_chunk
            inserted += len(rows)
            procesados += len(chunk)

            if rows:
                mov_values = [
                    (
                        usuario_id,
                        grupo_id,
                        "archivo",
                        r[0],
                        "CREAR_ARCHIVO",
                        None,
                        json_text({"id": r[0], "numero": r[1], "nombre": r[2], "caja_id": r[3], "tipo_doc": r[4]}),
                        None,
                    )
                    for r in rows
                ]
                insertar_movimientos_lote(cur, mov_values)

            conn.commit()
            set_import_job(
                job_id,
                processed_rows=min(procesados, total_rows),
                inserted=inserted,
                ignored=ignorados,
                invalid=invalidos,
                invalid_details=invalid_details,
                detail=f"Nuevos archivos: {inserted} | Ignorados: {ignorados} | Invalidos: {invalidos}",
            )

        final_status = "success" if not invalidos and not ignorados else "partial"
        detail = f"Nuevos archivos: {inserted} | Ignorados: {ignorados} | Invalidos: {invalidos}"
        set_import_job(
            job_id,
            status=final_status,
            message=(
                "Importacion finalizada correctamente."
                if final_status == "success"
                else "Importacion finalizada con observaciones."
            ),
            processed_rows=total_rows,
            inserted=inserted,
            ignored=ignorados,
            invalid=invalidos,
            invalid_details=invalid_details,
            error_code=None if final_status == "success" else 503,
            detail=detail,
            finished_at=datetime.utcnow().isoformat(),
            report_saved=True,
        )
        persist_import_report(get_import_job(job_id))
    except Exception:
        conn.rollback()
        app.logger.exception("Error importando Excel")
        set_import_job(
            job_id,
            status="failed",
            message=error_text(504),
            processed_rows=total_rows,
            error_code=504,
            finished_at=datetime.utcnow().isoformat(),
            report_saved=True,
        )
        persist_import_report(get_import_job(job_id))
    finally:
        cur.close()
        conn.close()
        try:
            os.remove(file_path)
        except Exception:
            pass


def importar_pdf_job(job_dir, file_names, grupo_id, usuario_id, job_id, client_ip=""):
    total_files = len(file_names or [])
    set_import_job(
        job_id,
        import_type="pdf",
        status="processing",
        message="Carga masiva de PDF en proceso.",
        total_rows=total_files,
        processed_rows=0,
        inserted=0,
        merged=0,
        ignored=0,
        invalid=0,
        invalid_details=[],
        error_code=None,
        finished_at=None,
        report_saved=False,
    )

    conn = get_db()
    cur = conn.cursor()
    nuevos = 0
    unidos = 0
    no_encontrados = 0
    invalidos = 0
    invalid_details = []
    procesados = 0

    try:
        for index, filename in enumerate(file_names, start=1):
            abs_path = os.path.join(job_dir, filename)
            display_name = obtener_nombre_original_pdf(filename)
            procesados = index

            if not filename.lower().endswith(".pdf"):
                invalidos += 1
                add_invalid_detail(invalid_details, display_name or filename, reason="extension invalida")
                set_import_job(
                    job_id,
                    processed_rows=procesados,
                    inserted=nuevos,
                    merged=unidos,
                    ignored=no_encontrados,
                    invalid=invalidos,
                    invalid_details=invalid_details,
                    detail=f"Nuevos: {nuevos} | Unidos: {unidos} | No encontrados: {no_encontrados} | Invalidos: {invalidos}",
                )
                continue

            numero = obtener_numero_desde_nombre_pdf(filename)
            if numero is None:
                invalidos += 1
                add_invalid_detail(invalid_details, display_name or filename, reason="nombre de archivo invalido")
                set_import_job(
                    job_id,
                    processed_rows=procesados,
                    inserted=nuevos,
                    merged=unidos,
                    ignored=no_encontrados,
                    invalid=invalidos,
                    invalid_details=invalid_details,
                    detail=f"Nuevos: {nuevos} | Unidos: {unidos} | No encontrados: {no_encontrados} | Invalidos: {invalidos}",
                )
                continue

            cur.execute(
                "SELECT id, pdf_path FROM archivos WHERE numero = %s AND grupo_id = %s",
                (numero, grupo_id)
            )
            row = cur.fetchone()
            if not row:
                no_encontrados += 1
                add_invalid_detail(invalid_details, display_name or filename, reason="documento no encontrado")
                set_import_job(
                    job_id,
                    processed_rows=procesados,
                    inserted=nuevos,
                    merged=unidos,
                    ignored=no_encontrados,
                    invalid=invalidos,
                    invalid_details=invalid_details,
                    detail=f"Nuevos: {nuevos} | Unidos: {unidos} | No encontrados: {no_encontrados} | Invalidos: {invalidos}",
                )
                continue

            archivo_id, pdf_old = row

            try:
                with open(abs_path, "rb") as fh:
                    pdf_bytes = fh.read()
            except Exception:
                invalidos += 1
                add_invalid_detail(invalid_details, display_name or filename, reason="no se pudo leer el archivo")
                set_import_job(
                    job_id,
                    processed_rows=procesados,
                    inserted=nuevos,
                    merged=unidos,
                    ignored=no_encontrados,
                    invalid=invalidos,
                    invalid_details=invalid_details,
                    detail=f"Nuevos: {nuevos} | Unidos: {unidos} | No encontrados: {no_encontrados} | Invalidos: {invalidos}",
                )
                continue

            class StoredUpload:
                def __init__(self, name, content):
                    self.filename = name
                    self.stream = BytesIO(content)

                def save(self, path):
                    self.stream.seek(0)
                    with open(path, "wb") as output:
                        output.write(self.stream.read())
                    self.stream.seek(0)

            file_obj = StoredUpload(display_name or filename, pdf_bytes)

            if pdf_old:
                pdf_name = unir_pdf_existente(pdf_old, file_obj, numero, grupo_id)
                if not pdf_name:
                    invalidos += 1
                    add_invalid_detail(invalid_details, display_name or filename, reason="no se pudo unir el PDF")
                    set_import_job(
                        job_id,
                        processed_rows=procesados,
                        inserted=nuevos,
                        merged=unidos,
                        ignored=no_encontrados,
                        invalid=invalidos,
                        invalid_details=invalid_details,
                        detail=f"Nuevos: {nuevos} | Unidos: {unidos} | No encontrados: {no_encontrados} | Invalidos: {invalidos}",
                    )
                    continue
                unidos += 1
            else:
                pdf_name = guardar_pdf(file_obj, numero, grupo_id)
                nuevos += 1

            cur.execute(
                "UPDATE archivos SET pdf_path = %s WHERE id = %s AND grupo_id = %s",
                (pdf_name, archivo_id, grupo_id)
            )

            if index % 25 == 0:
                conn.commit()

            set_import_job(
                job_id,
                processed_rows=procesados,
                inserted=nuevos,
                merged=unidos,
                ignored=no_encontrados,
                invalid=invalidos,
                invalid_details=invalid_details,
                detail=f"Nuevos: {nuevos} | Unidos: {unidos} | No encontrados: {no_encontrados} | Invalidos: {invalidos}",
            )

        conn.commit()

        registrar_movimiento(
            usuario_id,
            grupo_id,
            entidad="archivo",
            entidad_id=None,
            accion="CARGA_MASIVA_PDF",
            datos_despues={
                "procesados": procesados,
                "nuevos": nuevos,
                "unidos": unidos,
                "no_encontrados": no_encontrados,
                "invalidos": invalidos,
                "archivos_lote": total_files,
            }
        )

        registrar_log(
            usuario_id,
            f"CARGA_MASIVA_PDF lote={total_files} nuevos={nuevos} unidos={unidos} no_encontrados={no_encontrados} invalidos={invalidos}",
            client_ip,
            grupo_id
        )

        final_status = "success" if not no_encontrados and not invalidos else "partial"
        set_import_job(
            job_id,
            import_type="pdf",
            status=final_status,
            message=(
                "Carga masiva de PDF finalizada correctamente."
                if final_status == "success"
                else "Carga masiva de PDF finalizada con observaciones."
            ),
            processed_rows=total_files,
            inserted=nuevos,
            merged=unidos,
            ignored=no_encontrados,
            invalid=invalidos,
            invalid_details=invalid_details,
            error_code=None if final_status == "success" else 503,
            detail=f"Nuevos: {nuevos} | Unidos: {unidos} | No encontrados: {no_encontrados} | Invalidos: {invalidos}",
            finished_at=datetime.utcnow().isoformat(),
            report_saved=True,
        )
        persist_import_report(get_import_job(job_id))
    except Exception:
        conn.rollback()
        app.logger.exception("Error en carga masiva de PDF")
        set_import_job(
            job_id,
            import_type="pdf",
            status="failed",
            message=error_text(904, fallback="La carga masiva de PDF falló."),
            processed_rows=procesados,
            inserted=nuevos,
            merged=unidos,
            ignored=no_encontrados,
            invalid=invalidos,
            invalid_details=invalid_details,
            error_code=904,
            finished_at=datetime.utcnow().isoformat(),
            report_saved=True,
        )
        persist_import_report(get_import_job(job_id))
    finally:
        cur.close()
        conn.close()
        for filename in file_names or []:
            try:
                os.remove(os.path.join(job_dir, filename))
            except Exception:
                pass
        try:
            os.rmdir(job_dir)
        except Exception:
            pass


# ---------------- LOGIN ----------------
@app.route("/", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        usuario = request.form["usuario"]
        password = request.form["password"]

        info = verificar_usuario(usuario, password)  # (id, rol) o None
        if info:
            session["usuario"] = usuario
            session["usuario_id"] = info[0]
            session["rol"] = info[1]

            grupos = obtener_grupos_usuario(session["usuario_id"])
            if grupos:
                session["grupo_id"] = grupos[0][0]
            else:
                grupo_id = crear_grupo(f"Personal - {usuario}", creado_por=session["usuario_id"])
                agregar_usuario_a_grupo(
                    session["usuario_id"],
                    grupo_id,
                    puede_eliminar=True,
                    puede_editar=True
                )
                session["grupo_id"] = grupo_id

            # LOG: LOGIN
            registrar_log(
                session.get("usuario_id"),
                "LOGIN",
                request.remote_addr,
                session.get("grupo_id")
            )

            return redirect(url_for("inicio"))
        if usuario_existe(usuario):
            flash_error(2)
        else:
            flash_error(3)

    return render_template("login.html")


# ---------------- REGISTRO ----------------
@app.route("/register", methods=["GET", "POST"])
def register():
    if not login_requerido() or not admin_requerido():
        return "Acceso denegado", 403

    if request.method == "POST":
        usuario = request.form["usuario"]
        password = request.form["password"]

        try:
            user_id = crear_usuario(usuario, password, rol="cliente")
        except Exception as e:
            app.logger.exception("Error creando usuario")
            return render_template("register.html", error=f"Error al crear usuario: {e}")

        if user_id is False:
            return render_template("register.html", error="El usuario ya existe")

        grupo_id = crear_grupo(f"Personal - {usuario}", creado_por=user_id)
        agregar_usuario_a_grupo(user_id, grupo_id, puede_eliminar=True, puede_editar=True)

        return redirect(url_for("login"))

    # GET
    return render_template("register.html")



# ---------------- LOGOUT ----------------
@app.route("/logout")
def logout():
    # LOG: LOGOUT (antes de borrar sesiÃ³n)
    registrar_log(
        session.get("usuario_id"),
        "LOGOUT",
        request.remote_addr,
        session.get("grupo_id")
    )

    session.clear()
    return redirect(url_for("login"))


# ---------------- INICIO (BIENVENIDA) ----------------
@app.route("/inicio")
def inicio():
    if not login_requerido():
        return redirect(url_for("login"))

    grupo_id = obtener_grupo_id()
    if not grupo_id:
        return redirect(url_for("grupos"))

    conn = get_db()
    cur = conn.cursor()

    cur.execute(
        "SELECT COUNT(*) FROM archivos WHERE grupo_id = %s",
        (grupo_id,)
    )
    total_archivos = cur.fetchone()[0]

    cur.execute(
        "SELECT COUNT(*) FROM cajas WHERE grupo_id = %s AND is_pendiente = FALSE",
        (grupo_id,)
    )
    total_cajas = cur.fetchone()[0]

    cur.close()
    conn.close()

    return render_template(
        "inicio.html",
        total_archivos=total_archivos,
        total_cajas=total_cajas
    )


# ---------------- GRUPOS ----------------
@app.route("/grupos", methods=["GET", "POST"])
def grupos():
    if not login_requerido():
        return redirect(url_for("login"))

    if request.method == "POST" and not admin_requerido():
        accion = request.form.get("accion")
        if accion != "actualizar_permiso":
            flash_error(200)
            return redirect(url_for("grupos"))

        if not supervisor_requerido():
            flash_error(201)
            return redirect(url_for("grupos"))

        grupo_id = int(request.form.get("grupo_id"))
        usuario_id = int(request.form.get("usuario_id"))
        puede_eliminar = request.form.get("puede_eliminar") == "1"
        puede_editar = request.form.get("puede_editar") == "1"

        grupos_usuario = {g[0] for g in obtener_grupos_usuario(session.get("usuario_id"))}
        if grupo_id not in grupos_usuario:
            flash_error(202)
            return redirect(url_for("grupos"))

        agregar_usuario_a_grupo(usuario_id, grupo_id, puede_eliminar, puede_editar)
        flash("Permisos actualizados.", "success")
        return redirect(url_for("grupos"))

    if admin_requerido():
        if request.method == "POST":
            accion = request.form.get("accion")

            if accion == "crear_grupo":
                nombre = request.form.get("nombre", "").strip()
                if nombre:
                    crear_grupo(nombre, creado_por=session.get("usuario_id"))
                    flash("Grupo creado correctamente.", "success")
                else:
                    flash_error(300)

            if accion == "agregar_usuario":
                grupo_id = int(request.form.get("grupo_id"))
                usuario = request.form.get("usuario", "").strip()
                puede_eliminar = request.form.get("puede_eliminar") == "1"
                puede_editar = request.form.get("puede_editar") == "1"

                user_row = buscar_usuario_por_nombre(usuario)
                if not user_row:
                    flash_error(301)
                else:
                    agregar_usuario_a_grupo(user_row[0], grupo_id, puede_eliminar, puede_editar)
                    eliminar_grupo_personal(user_row[0], grupo_id)
                    flash("Usuario agregado al grupo.", "success")

            if accion == "actualizar_permiso":
                grupo_id = int(request.form.get("grupo_id"))
                usuario_id = int(request.form.get("usuario_id"))
                puede_eliminar = request.form.get("puede_eliminar") == "1"
                puede_editar = request.form.get("puede_editar") == "1"

                agregar_usuario_a_grupo(usuario_id, grupo_id, puede_eliminar, puede_editar)
                flash("Permisos actualizados.", "success")

            if accion == "quitar_usuario":
                grupo_id = int(request.form.get("grupo_id"))
                usuario_id = int(request.form.get("usuario_id"))
                quitar_usuario_de_grupo(usuario_id, grupo_id)
                flash("Usuario removido del grupo.", "success")

            if accion == "eliminar_grupo":
                grupo_id = int(request.form.get("grupo_id"))
                ok = archivar_grupo(grupo_id, session.get("usuario_id"))
                if ok:
                    flash("Grupo archivado y eliminado correctamente.", "success")
                else:
                    flash_error(307)

            if accion == "crear_usuario":
                usuario = request.form.get("usuario", "").strip()
                password = request.form.get("password", "").strip()
                rol = request.form.get("rol", "cliente").strip()
                if not usuario or not password:
                    flash_error(303)
                else:
                    try:
                        user_id = crear_usuario(usuario, password, rol=rol)
                        if user_id is False:
                            flash_error(302)
                        else:
                            grupo_id = crear_grupo(f"Personal - {usuario}", creado_por=user_id)
                            agregar_usuario_a_grupo(user_id, grupo_id, puede_eliminar=True, puede_editar=True)
                            flash("Usuario creado correctamente.", "success")
                    except Exception as e:
                        app.logger.exception("Error creando usuario desde grupos")
                        flash_error(304, detail=str(e))

            if accion == "actualizar_usuario":
                usuario_id = int(request.form.get("usuario_id"))
                nuevo_usuario = request.form.get("usuario", "").strip()
                nuevo_rol = request.form.get("rol", "").strip()
                nueva_password = request.form.get("password", "").strip()
                nuevo_grupo_id = int(request.form.get("grupo_id"))
                return_to = request.form.get("return_to", "")

                conn = get_db()
                cur = conn.cursor()

                try:
                    if nuevo_usuario:
                        cur.execute(
                            "UPDATE usuarios SET usuario = %s WHERE id = %s",
                            (nuevo_usuario, usuario_id)
                        )

                    if nuevo_rol:
                        cur.execute(
                            "UPDATE usuarios SET rol = %s WHERE id = %s",
                            (nuevo_rol, usuario_id)
                        )

                        cur.execute(
                            """
                            UPDATE usuariosempresas
                            SET eliminar = %s
                            WHERE usuarioid = %s
                            """,
                            ((nuevo_rol in ("admin", "supervisor")), usuario_id)
                        )

                    if nueva_password:
                        hash_pw = generate_password_hash(nueva_password)
                        cur.execute(
                            "UPDATE usuarios SET [contraseña] = %s WHERE id = %s",
                            (hash_pw, usuario_id)
                        )

                    cur.execute(
                        """
                        SELECT g.id, g.nombre
                        FROM usuariosempresas ug
                        JOIN empresas g ON g.id = ug.empresa
                        WHERE ug.usuarioid = %s
                        ORDER BY g.id
                        LIMIT 1
                        """,
                        (usuario_id,)
                    )
                    current = cur.fetchone()
                    current_group_id = current[0] if current else None
                    current_group_name = current[1] if current else None

                    if nuevo_grupo_id == 0:
                        cur.execute(
                            "DELETE FROM usuariosempresas WHERE usuarioid = %s",
                            (usuario_id,)
                        )
                        cur.execute(
                            "DELETE FROM usuariosregistrados WHERE usuarioid = %s",
                            (usuario_id,)
                        )
                    elif current_group_id != nuevo_grupo_id:
                        cur.execute(
                            "DELETE FROM usuariosempresas WHERE usuarioid = %s",
                            (usuario_id,)
                        )
                        cur.execute(
                            "DELETE FROM usuariosregistrados WHERE usuarioid = %s",
                            (usuario_id,)
                        )
                        agregar_usuario_a_grupo(
                            usuario_id,
                            nuevo_grupo_id,
                            puede_eliminar=(nuevo_rol in ("admin", "supervisor")),
                            puede_editar=True
                        )
                        if current_group_name and current_group_name.startswith("Personal - "):
                            eliminar_grupo_personal(usuario_id, nuevo_grupo_id)

                    conn.commit()
                    sincronizar_registros_usuario(usuario_id)
                    flash("Usuario actualizado correctamente.", "success")
                except Exception as e:
                    conn.rollback()
                    flash_error(305, detail=str(e))
                finally:
                    cur.close()
                    conn.close()
                if return_to == "admin_logs_usuarios":
                    return redirect(url_for("admin_logs", tab="usuarios"))

            if accion == "eliminar_usuario":
                usuario_id = int(request.form.get("usuario_id"))
                return_to = request.form.get("return_to", "")
                if usuario_id == session.get("usuario_id"):
                    flash_error(207)
                    return redirect(url_for("grupos"))

                conn = get_db()
                cur = conn.cursor()
                try:
                    # Desvincular logs/movimientos para evitar FK al eliminar usuario
                    cur.execute("UPDATE logs SET usuario_id = NULL WHERE usuario_id = %s", (usuario_id,))
                    cur.execute("UPDATE movimientos SET usuario_id = NULL WHERE usuario_id = %s", (usuario_id,))
                    # Desvincular creador en grupos/cajas/archivos
                    cur.execute("UPDATE grupos SET creado_por = NULL WHERE creado_por = %s", (usuario_id,))
                    cur.execute("UPDATE cajas SET creado_por = NULL WHERE creado_por = %s", (usuario_id,))
                    cur.execute("UPDATE archivos SET creado_por = NULL WHERE creado_por = %s", (usuario_id,))
                    cur.execute("DELETE FROM usuariosempresas WHERE usuarioid = %s", (usuario_id,))
                    cur.execute("DELETE FROM usuariosregistrados WHERE usuarioid = %s", (usuario_id,))
                    cur.execute("DELETE FROM usuarios WHERE id = %s", (usuario_id,))
                    conn.commit()
                except Exception as e:
                    conn.rollback()
                    flash_error(306, detail=str(e))
                    cur.close()
                    conn.close()
                    return redirect(url_for("grupos"))
                cur.close()
                conn.close()
                flash("Usuario eliminado correctamente.", "success")
                if return_to == "admin_logs_usuarios":
                    return redirect(url_for("admin_logs", tab="usuarios"))

            return redirect(url_for("grupos"))

        grupos_data = obtener_todos_grupos()
        grupos_info = []
        for g in grupos_data:
            miembros = obtener_miembros_grupo(g[0])
            grupos_info.append((g[0], g[1], miembros))

        return render_template("grupos.html", grupos=grupos_info, es_admin=True)

    if usuario_requerido():
        return redirect(url_for("archivo"))

    grupos_data = obtener_grupos_usuario(session.get("usuario_id"))
    grupos_info = []
    conn = get_db()
    cur = conn.cursor()
    log_cols = set()
    try:
        cur.execute(
            """
            SELECT column_name
            FROM information_schema.columns
            WHERE table_name = 'logs'
            """
        )
        log_cols = {r[0] for r in cur.fetchall()}
    except Exception:
        conn.rollback()
        log_cols = set()

    if "fecha" in log_cols:
        fecha_expr = "MAX(l.fecha)"
    elif "creado_en" in log_cols:
        fecha_expr = "MAX(l.creado_en)"
    elif "created_at" in log_cols:
        fecha_expr = "MAX(l.created_at)"
    else:
        fecha_expr = "NULL"

    for g in grupos_data:
        cur.execute(
            f"""
            SELECT
                u.id,
                u.usuario,
                ug.eliminar,
                ug.editar,
                {fecha_expr} AS ultima_conexion
            FROM usuariosempresas ug
            JOIN usuarios u ON u.id = ug.usuarioid
            LEFT JOIN auditoria l ON l.usuarioid = u.id AND l.accion = 'LOGIN'
            WHERE ug.empresa = %s
            GROUP BY u.id, u.usuario, ug.eliminar, ug.editar
            ORDER BY u.usuario
            """,
            (g[0],)
        )
        miembros = cur.fetchall()
        cur.execute(
            """
            WITH ranked AS (
                SELECT id, ROW_NUMBER() OVER (ORDER BY rango_min, id) AS caja_visible
                FROM cajas
                WHERE grupo_id = %s AND is_pendiente = FALSE
            )
            SELECT
                c.id,
                CASE WHEN c.is_pendiente = 1 THEN 0 ELSE r.caja_visible END AS caja_visible
            FROM cajas c
            LEFT JOIN ranked r ON r.id = c.id
            WHERE c.grupo_id = %s
            """,
            (g[0], g[0])
        )
        caja_map = {row[0]: row[1] for row in cur.fetchall()}

        def _json_or_none(val):
            if val is None:
                return None
            if isinstance(val, (dict, list)):
                return val
            try:
                return json.loads(val)
            except Exception:
                return None

        def _fmt_num(val):
            try:
                return f"{int(val):,}".replace(",", ".")
            except Exception:
                return val

        movimientos_por_usuario = {}
        for m in miembros:
            cur.execute(
                """
                SELECT
                    m.id,
                    m.fecha,
                    m.item,
                    m.accion,
                    m.antes,
                    m.despues
                FROM movimientos m
                WHERE m.usuarioid = %s AND m.empresa = %s
                ORDER BY m.fecha DESC
                LIMIT 20
                """,
                (m[0], g[0])
            )
            rows = cur.fetchall()
            movs = []
            for mid, fecha, item_raw, accion, antes_raw, despues_raw in rows:
                antes = _json_or_none(antes_raw) or {}
                despues = _json_or_none(despues_raw) or {}
                entidad, entidad_id = parse_item_movimiento(item_raw)

                doc_num_raw = antes.get("numero") or despues.get("numero")
                doc_num = _fmt_num(doc_num_raw)
                caja_old = antes.get("caja_id")
                caja_new = despues.get("caja_id")
                caja_old_vis = caja_map.get(caja_old, caja_old)
                caja_new_vis = caja_map.get(caja_new, caja_new)
                caja_vis = caja_map.get(entidad_id, entidad_id)

                if entidad == "archivo" and accion == "ARCHIVO_MOVER":
                    texto = f"Se movio el documento {doc_num} de la caja {caja_old_vis} a la caja {caja_new_vis}"
                elif entidad == "archivo" and accion == "CREAR_ARCHIVO":
                    texto = f"Se creo el documento {doc_num}"
                elif entidad == "archivo" and accion == "MODIFICAR_ARCHIVO":
                    texto = f"Se modifico el documento {doc_num}"
                elif entidad == "archivo" and accion == "ELIMINAR_ARCHIVO":
                    texto = f"Se elimino el documento {doc_num}"
                elif entidad == "archivo" and accion == "CARGA_MASIVA_PDF":
                    nuevos = despues.get("nuevos", 0)
                    unidos = despues.get("unidos", 0)
                    no_encontrados = despues.get("no_encontrados", 0)
                    invalidos = despues.get("invalidos", 0)
                    lote = despues.get("archivos_lote", 0)
                    texto = (
                        f"CARGA_MASIVA_PDF: lote {lote}, {nuevos} nuevos, "
                        f"{unidos} unidos, {no_encontrados} no encontrados, {invalidos} invalidos"
                    )
                elif entidad == "caja" and accion == "CREAR_CAJA":
                    texto = f"Se creo la caja numero {caja_vis}"
                elif entidad == "caja" and accion == "MODIFICAR_CAJA":
                    rango_old = f"{_fmt_num(antes.get('rango_min'))}-{_fmt_num(antes.get('rango_max'))}"
                    rango_new = f"{_fmt_num(despues.get('rango_min'))}-{_fmt_num(despues.get('rango_max'))}"
                    reasignados = despues.get("reasignados", 0)
                    texto = f"MODIFICAR_CAJA masivo: {reasignados} archivos reasignados ({rango_old}) a ({rango_new}) en la caja {caja_vis}"
                elif entidad == "caja" and accion == "ELIMINAR_CAJA":
                    reasignados = despues.get("reasignados", 0)
                    rango_old = f"{_fmt_num(antes.get('rango_min'))}-{_fmt_num(antes.get('rango_max'))}"
                    texto = f"ELIMINAR_CAJA masivo: {reasignados} archivos reasignados desde la caja {caja_vis} ({rango_old})"
                elif entidad == "caja" and accion == "ELIMINAR_ARCHIVOS_CAJA":
                    total_archivos = despues.get("archivos_eliminados", antes.get("total_archivos", 0))
                    total_pdfs = despues.get("pdfs_eliminados", 0)
                    texto = f"ELIMINAR_ARCHIVOS_CAJA permanente: {total_archivos} documentos y {total_pdfs} PDF(s) eliminados de la caja {caja_vis}"
                else:
                    texto = f"{accion}"

                movs.append((fecha, texto, mid))
            movimientos_por_usuario[m[0]] = movs
        grupos_info.append((g[0], g[1], g[2], g[3], miembros))
    cur.close()
    conn.close()
    empresa_nombre = grupos_info[0][1] if grupos_info else "Empresa"
    return render_template(
        "grupos.html",
        grupos=grupos_info,
        es_admin=False,
        empresa_nombre=empresa_nombre,
        movimientos_por_usuario=movimientos_por_usuario
    )


@app.context_processor
def inject_grupo_actual():
    nombre = None
    grupo_id = session.get("grupo_id")
    if grupo_id:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("SELECT nombre FROM grupos WHERE id = %s", (grupo_id,))
        row = cur.fetchone()
        cur.close()
        conn.close()
        nombre = row[0] if row else None
        if nombre and nombre.startswith("Personal - "):
            nombre = "Personal"
    return {"grupo_actual": nombre}


@app.route("/grupos/seleccionar/<int:grupo_id>")
def seleccionar_grupo(grupo_id):
    if not login_requerido():
        return redirect(url_for("login"))

    if not admin_requerido():
        grupos_usuario = {g[0] for g in obtener_grupos_usuario(session.get("usuario_id"))}
        if grupo_id not in grupos_usuario:
            return "Acceso denegado", 403

    session["grupo_id"] = grupo_id
    return redirect(url_for("archivo"))


# ---------------- CAJAS ----------------
@app.route("/cajas", methods=["GET", "POST"])
def cajas():
    if not login_requerido():
        return redirect(url_for("login"))

    grupo_id = obtener_grupo_id()
    if not grupo_id:
        return redirect(url_for("grupos"))

    # Asegurar que siempre exista la caja pendiente
    asegurar_caja_sin_asignar(grupo_id)

    if request.method == "POST":
        accion = request.form.get("accion")

        # ======================
        # ELIMINAR CAJA
        # ======================
        if accion == "eliminar":
            caja_id = int(request.form["caja_id"])

            if not puede_eliminar_cajas():
                flash_error(355, fallback="Solo el supervisor o admin puede eliminar cajas.")
                return redirect(url_for("cajas"))

            caja_eliminada = eliminar_caja(caja_id, grupo_id, session.get("usuario_id"))

            if caja_eliminada:
                registrar_movimiento(
                    session.get("usuario_id"),
                    grupo_id,
                    entidad="caja",
                    entidad_id=caja_id,
                    accion="ELIMINAR_CAJA",
                    datos_antes={
                        "rango_min": caja_eliminada["rango_min"],
                        "rango_max": caja_eliminada["rango_max"],
                    },
                    datos_despues={
                        "reasignados": caja_eliminada["reasignados"],
                        "modo": "masivo",
                    },
                )

            registrar_log(
                session.get("usuario_id"),
                f"ELIMINAR_CAJA_MASIVO caja={caja_id} movidos={caja_eliminada['reasignados'] if caja_eliminada else 0}",
                request.remote_addr,
                grupo_id
            )

            flash("Caja eliminada correctamente.", "success")
            return redirect(url_for("cajas"))

        # ======================
        # MODIFICAR CAJA (modal)
        # ======================
        if accion == "modificar":
            caja_id = int(request.form["caja_id"])
            nuevo_min = int(request.form["rango_min"])
            nuevo_max = int(request.form["rango_max"])

            if not puede_crear_modificar_cajas():
                flash_error(204)
                return redirect(url_for("cajas"))

            caja_antes = modificar_caja(caja_id, nuevo_min, nuevo_max, grupo_id, session.get("usuario_id"))

            movidos_fuera = reubicar_archivos_de_caja(caja_id, grupo_id, session.get("usuario_id"))
            movidos_dentro = reubicar_archivos_pendientes_por_nueva_caja(
                caja_id,
                grupo_id,
                session.get("usuario_id")
            )

            total_movidos = movidos_fuera + movidos_dentro

            if total_movidos > 0:
                flash(
                    f"Se reasignaron automaticamente {total_movidos} archivo(s) segun el nuevo rango.",
                    "success"
                )
            else:
                flash(
                    "El rango se actualizo. No fue necesario reasignar archivos.",
                    "info"
                )

            if caja_antes:
                registrar_movimiento(
                    session.get("usuario_id"),
                    grupo_id,
                    entidad="caja",
                    entidad_id=caja_id,
                    accion="MODIFICAR_CAJA",
                    datos_antes={
                        "rango_min": caja_antes["rango_min"],
                        "rango_max": caja_antes["rango_max"],
                    },
                    datos_despues={
                        "rango_min": caja_antes["nuevo_min"],
                        "rango_max": caja_antes["nuevo_max"],
                        "reasignados": total_movidos,
                        "modo": "masivo",
                    },
                )

            registrar_log(
                session.get("usuario_id"),
                f"MODIFICAR_CAJA_MASIVO caja={caja_id} movidos={total_movidos}",
                request.remote_addr,
                grupo_id
            )

            return redirect(url_for("cajas"))

        # ======================
        # CREAR CAJA
        # ======================
        if accion == "crear":
            rmin = int(request.form["rango_min"])
            rmax = int(request.form["rango_max"])

            if not puede_crear_modificar_cajas():
                flash_error(205)
                return redirect(url_for("cajas"))

            nueva_caja_id = crear_caja(
                rmin,
                rmax,
                grupo_id,
                creado_por=session.get("usuario_id")
            )

            # Reubicar archivos pendientes (Caja 0) si encajan en la nueva caja
            movidos = reubicar_archivos_pendientes_por_nueva_caja(
                nueva_caja_id,
                grupo_id,
                session.get("usuario_id")
            )

            registrar_log(
                session.get("usuario_id"),
                f"CREAR_CAJA caja_id={nueva_caja_id} rango={rmin}-{rmax} movidos_desde_caja0={movidos}",
                request.remote_addr,
                grupo_id
            )

            if movidos > 0:
                flash(f"Caja creada. Se movieron {movidos} archivo(s) desde Caja 0.", "success")
            else:
                flash("Caja creada correctamente.", "success")

            return redirect(url_for("cajas"))

        # Si llega algo raro:
        flash_error(200)
        return redirect(url_for("cajas"))

    # ======================
    # LISTADO DE CAJAS
    # ======================
    conn = get_db()
    cur = conn.cursor()

    cur.execute(
        """
        WITH ranked AS (
            SELECT
                id,
                rango_min,
                rango_max,
                ROW_NUMBER() OVER (ORDER BY rango_min, id) AS caja_visible
            FROM cajas
            WHERE grupo_id = %s AND is_pendiente = FALSE
        )
        SELECT
            c.id AS caja_id,
            COUNT(a.id) AS total_archivos,
            c.rango_min,
            c.rango_max,
            CASE
                WHEN c.is_pendiente = 1 THEN 0
                ELSE r.caja_visible
            END AS caja_visible,
            c.is_pendiente
        FROM cajas c
        LEFT JOIN archivos a ON a.caja_id = c.id AND a.grupo_id = %s
        LEFT JOIN ranked r ON r.id = c.id
        WHERE c.grupo_id = %s
        GROUP BY c.id, c.rango_min, c.rango_max, c.is_pendiente, r.caja_visible
        ORDER BY
            CASE WHEN c.is_pendiente = 1 THEN 1 ELSE 0 END,
            caja_visible
        """,
        (grupo_id, grupo_id, grupo_id)
    )

    cajas_data = cur.fetchall()
    cur.close()
    conn.close()

    # Ocultar Caja 0 si estÃ¡ vacÃ­a
    cajas_filtradas = []
    for c in cajas_data:
        if c[5] and c[1] == 0:
            continue
        cajas_filtradas.append(c)

    return render_template("cajas.html", cajas=cajas_filtradas)



# ---------------- ARCHIVOS ----------------
@app.route("/archivos", methods=["GET", "POST"])
def archivos():
    # Pestaña antigua: redirigir a Archivo
    return redirect(url_for("archivo"))

@app.route("/archivos_legacy", methods=["GET", "POST"])
def archivos_legacy():
    if not login_requerido():
        return redirect(url_for("login"))

    grupo_id = obtener_grupo_id()
    if not grupo_id:
        return redirect(url_for("grupos"))

    # =========================
    # POST: agregar / modificar fila / eliminar fila
    # =========================
    if request.method == "POST":
        accion = request.form["accion"]


        # âœ… AGREGAR (formulario de arriba)
        if accion == "agregar":
            numero = int(request.form["numero"])
            tipo_doc = normalizar_tipo_doc(request.form.get("tipo_doc", ""))
            if not es_tipo_doc_valido(tipo_doc):
                flash_error(400)
                return redirect(url_for("archivos"))
            nombre = request.form.get("nombre", "").strip()
            if not nombre:
                nombre = f"Documento {numero}"
            nombre = normalizar_nombre(nombre)

            agregar_archivo(
                numero,
                nombre,
                grupo_id,
                creado_por=session.get("usuario_id"),
                tipo_doc=tipo_doc
            )

            # Log con info real
            conn = get_db()
            cur = conn.cursor()
            cur.execute(
                "SELECT caja_id, nombre FROM archivos WHERE numero = %s AND grupo_id = %s",
                (numero, grupo_id)
            )
            info = cur.fetchone()
            caja_id = info[0] if info else 0
            nombre_final = info[1] if info else nombre
            cur.close()
            conn.close()

            registrar_log(
                session.get("usuario_id"),
                f"ARCHIVO|tipo=REGISTRO|numero_new={numero}|nombre_new={nombre_final}|caja={caja_id}",
                request.remote_addr,
                grupo_id
            )


            return redirect(url_for("archivos"))

       # âœ… MODIFICAR DESDE FILA (popup)
        if accion == "modificar_fila":
            numero_old = int(request.form["numero_old"])
            numero_new = int(request.form["numero_new"])
            nombre_new = request.form.get("nombre_new", "").strip()
            tipo_doc_new = normalizar_tipo_doc(request.form.get("tipo_doc_new", ""))
            if not es_tipo_doc_valido(tipo_doc_new):
                flash_error(400)
                return redirect(url_for("archivos"))
            nombre_new = normalizar_nombre(nombre_new)
            nombre_new = normalizar_nombre(nombre_new)

            conn = get_db()
            cur = conn.cursor()

            # Antes:
            cur.execute(
                "SELECT id, caja_id, numero, nombre, pdf_path, tipo_doc FROM archivos WHERE numero = %s AND grupo_id = %s",
                (numero_old, grupo_id)
            )
            antes = cur.fetchone()
            archivo_id = antes[0] if antes else None
            caja_id = antes[1] if antes else 0
            numero_viejo = antes[2] if antes else numero_old
            nombre_viejo = antes[3] if antes else ""
            pdf_old = antes[4] if antes else None
            tipo_doc_old = antes[5] if antes else "CC"

            # ... haces el UPDATE ...

            registrar_log(
                session.get("usuario_id"),
                f"ARCHIVO|tipo=MODIFICACION|numero_old={numero_viejo}|numero_new={numero_new}|nombre_old={nombre_viejo}|nombre_new={nombre_new}|caja={caja_id}",
                request.remote_addr,
                grupo_id
            )

            cur.execute("""
                UPDATE archivos
                SET numero = %s, nombre = %s, tipo_doc = %s
                WHERE numero = %s AND grupo_id = %s
            """, (numero_new, nombre_new, tipo_doc_new, numero_old, grupo_id))

            conn.commit()
            cur.close()
            conn.close()

            registrar_log(
                session.get("usuario_id"),
                f"MODIFICAR_ARCHIVO | caja_id={caja_id} | numero={numero_new} | nombre={nombre_new}",
                request.remote_addr,
                grupo_id
            )

            if archivo_id:
                registrar_movimiento(
                    session.get("usuario_id"),
                    grupo_id,
                    entidad="archivo",
                    entidad_id=archivo_id,
                    accion="MODIFICAR_ARCHIVO",
                    datos_antes={
                        "numero": numero_viejo,
                        "nombre": nombre_viejo,
                        "caja_id": caja_id,
                        "pdf_path": pdf_old,
                        "tipo_doc": tipo_doc_old,
                    },
                    datos_despues={
                        "numero": numero_new,
                        "nombre": nombre_new,
                        "caja_id": caja_id,
                        "pdf_path": pdf_old,
                        "tipo_doc": tipo_doc_new,
                    },
                )

            return redirect(url_for("archivos"))


        # âœ… ELIMINAR DESDE FILA
        if accion == "eliminar_fila":
            numero = int(request.form["numero"])

            if not supervisor_requerido() and not usuario_puede_eliminar(session.get("usuario_id"), grupo_id):
                flash_error(203)
                return redirect(url_for("archivos"))

            conn = get_db()
            cur = conn.cursor()

            # guardar info antes de borrar
            cur.execute(
                "SELECT caja_id, nombre FROM archivos WHERE numero = %s AND grupo_id = %s",
                (numero, grupo_id)
            )
            antes = cur.fetchone()
            caja_id = antes[0] if antes else 0
            nombre_final = antes[1] if antes else ""

            cur.close()
            conn.close()

            eliminar_archivo(numero, grupo_id, session.get("usuario_id"))

            registrar_log(
                session.get("usuario_id"),
                f"ARCHIVO|tipo=ELIMINACION|numero_old={numero}|nombre_old={nombre_final}|caja={caja_id}",
                request.remote_addr,
                grupo_id
            )

            return redirect(url_for("archivos"))

        # Si llega una acciÃ³n desconocida:
        return redirect(url_for("archivos"))

    # =========================
    # GET: listado + bÃºsqueda + Ãºltimos movimientos + modo ediciÃ³n
    # =========================
    conn = get_db()
    cur = conn.cursor()

    # Listado normal (caja visible 1..N y caja 0)
    cur.execute(
          """
          WITH ranked AS (
              SELECT
                  id,
                  ROW_NUMBER() OVER (ORDER BY rango_min, id) AS caja_visible
              FROM cajas
              WHERE grupo_id = %s AND is_pendiente = FALSE
          )
          SELECT
              CASE WHEN c.is_pendiente = 1 THEN 0 ELSE r.caja_visible END AS caja,
              a.tipo_doc AS tipo_doc,
              a.numero AS documento,
              a.nombre AS nombre
          FROM archivos a
          JOIN cajas c ON a.caja_id = c.id
          LEFT JOIN ranked r ON r.id = c.id
          WHERE a.grupo_id = %s
          ORDER BY caja, a.numero
        """,
        (grupo_id, grupo_id)
    )

    datos = cur.fetchall()

    # Buscador
    resultado = None
    buscar = request.args.get("buscar", "").strip()
    if buscar:
        buscar_clean = buscar.replace(".", "")
        if buscar_clean.isdigit():
            doc = int(buscar_clean)
            if "." in buscar:
                cur.execute("""
                    WITH ranked AS (
                    SELECT id, ROW_NUMBER() OVER (ORDER BY rango_min, id) AS caja_visible
                    FROM cajas
                    WHERE grupo_id = %s AND is_pendiente = FALSE
                    )
                      SELECT
                      c.id AS caja_id,
                      CASE WHEN c.is_pendiente = 1 THEN 0 ELSE r.caja_visible END AS caja_num,
                      a.tipo_doc AS tipo_doc,
                      a.numero AS documento,
                      a.nombre AS nombre,
                      a.pdf_path
                      FROM archivos a
                    JOIN cajas c ON c.id = a.caja_id
                    LEFT JOIN ranked r ON r.id = c.id
                    WHERE a.grupo_id = %s
                      AND (
                        a.numero = %s
                        OR a.nombre ILIKE %s
                      )
                    ORDER BY a.numero
                """, (grupo_id, grupo_id, doc, f"%{buscar}%"))
            else:
                cur.execute("""
                    WITH ranked AS (
                    SELECT id, ROW_NUMBER() OVER (ORDER BY rango_min, id) AS caja_visible
                    FROM cajas
                    WHERE grupo_id = %s AND is_pendiente = FALSE
                    )
                      SELECT
                      c.id AS caja_id,
                      CASE WHEN c.is_pendiente = 1 THEN 0 ELSE r.caja_visible END AS caja_num,
                      a.tipo_doc AS tipo_doc,
                      a.numero AS documento,
                      a.nombre AS nombre,
                      a.pdf_path
                      FROM archivos a
                    JOIN cajas c ON c.id = a.caja_id
                    LEFT JOIN ranked r ON r.id = c.id
                    WHERE a.grupo_id = %s
                      AND (
                        a.numero = %s
                        OR replace(a.nombre, '.', '') = %s
                      )
                    ORDER BY a.numero
                """, (grupo_id, grupo_id, doc, buscar_clean))

            rows = cur.fetchall()
            if rows:
                resultado = rows_to_builtin(rows)
            else:
                resultado = ("no",)
        else:
            cur.execute("""
                WITH ranked AS (
                SELECT id, ROW_NUMBER() OVER (ORDER BY rango_min, id) AS caja_visible
                FROM cajas
                WHERE grupo_id = %s AND is_pendiente = FALSE
                )
                  SELECT
                  c.id AS caja_id,
                  CASE WHEN c.is_pendiente = 1 THEN 0 ELSE r.caja_visible END AS caja_num,
                  a.tipo_doc AS tipo_doc,
                  a.numero AS documento,
                  a.nombre AS nombre,
                  a.pdf_path
                  FROM archivos a
                JOIN cajas c ON c.id = a.caja_id
                LEFT JOIN ranked r ON r.id = c.id
                WHERE a.nombre ILIKE %s AND a.grupo_id = %s
                ORDER BY a.numero
            """, (grupo_id, f"%{buscar}%", grupo_id))
            rows = cur.fetchall()
            if rows:
                resultado = rows_to_builtin(rows)
            else:
                resultado = ("no",)

    # Modo ediciÃ³n (quÃ© documento estÃ¡ en ediciÃ³n)
    edit_num = request.args.get("edit", "").strip()
    edit_num = int(edit_num) if edit_num.isdigit() else None

    # Ultimos movimientos (20)
    cur.execute("""
        WITH ranked AS (
            SELECT id, ROW_NUMBER() OVER (ORDER BY rango_min, id) AS caja_visible
            FROM cajas
            WHERE grupo_id = %s AND is_pendiente = FALSE
        )
        SELECT
            c.id,
            CASE WHEN c.is_pendiente = 1 THEN 0 ELSE r.caja_visible END AS caja_visible
        FROM cajas c
        LEFT JOIN ranked r ON r.id = c.id
        WHERE c.grupo_id = %s
    """, (grupo_id, grupo_id))
    caja_map = {row[0]: row[1] for row in cur.fetchall()}

    cur.execute("""
        SELECT
            m.id,
            m.fecha,
            m.entidad,
            m.entidad_id,
            m.accion,
            m.datos_antes,
            m.datos_despues,
            a.numero
        FROM movimientos m
        LEFT JOIN archivos a ON a.id = m.entidad_id AND m.entidad = 'archivo'
        WHERE m.grupo_id = %s
        ORDER BY m.fecha DESC
        LIMIT 20
    """, (grupo_id,))
    mov_rows = cur.fetchall()

    def _json_or_none(val):
        if val is None:
            return None
        if isinstance(val, (dict, list)):
            return val
        try:
            return json.loads(val)
        except Exception:
            return None

    def _fmt_num(val):
        try:
            return f"{int(val):,}".replace(",", ".")
        except Exception:
            return val

    movimientos = []
    for mid, fecha, entidad, entidad_id, accion, antes_raw, despues_raw, numero in mov_rows:
        antes = _json_or_none(antes_raw) or {}
        despues = _json_or_none(despues_raw) or {}

        doc_num_raw = numero or antes.get("numero") or despues.get("numero")
        doc_num = _fmt_num(doc_num_raw)
        caja_old = antes.get("caja_id")
        caja_new = despues.get("caja_id")

        caja_old_vis = caja_map.get(caja_old, caja_old)
        caja_new_vis = caja_map.get(caja_new, caja_new)
        caja_vis = caja_map.get(entidad_id, entidad_id)

        if entidad == "archivo" and accion == "ARCHIVO_MOVER":
            texto = f"Se movio el documento {doc_num} de la caja {caja_old_vis} a la caja {caja_new_vis}"
        elif entidad == "archivo" and accion == "CREAR_ARCHIVO":
            texto = f"Se creo el documento {doc_num}"
        elif entidad == "archivo" and accion == "MODIFICAR_ARCHIVO":
            texto = f"Se modifico el documento {doc_num}"
        elif entidad == "archivo" and accion == "ELIMINAR_ARCHIVO":
            texto = f"Se elimino el documento {doc_num}"
        elif entidad == "archivo" and accion == "CARGA_MASIVA_PDF":
            nuevos = despues.get("nuevos", 0)
            unidos = despues.get("unidos", 0)
            no_encontrados = despues.get("no_encontrados", 0)
            invalidos = despues.get("invalidos", 0)
            lote = despues.get("archivos_lote", 0)
            texto = (
                f"CARGA_MASIVA_PDF: lote {lote}, {nuevos} nuevos, "
                f"{unidos} unidos, {no_encontrados} no encontrados, {invalidos} invalidos"
            )
        elif entidad == "caja" and accion == "CREAR_CAJA":
            texto = f"Se creo la caja numero {caja_vis}"
        elif entidad == "caja" and accion == "MODIFICAR_CAJA":
            rango_old = f"{_fmt_num(antes.get('rango_min'))}-{_fmt_num(antes.get('rango_max'))}"
            rango_new = f"{_fmt_num(despues.get('rango_min'))}-{_fmt_num(despues.get('rango_max'))}"
            reasignados = despues.get("reasignados", 0)
            texto = f"MODIFICAR_CAJA masivo: {reasignados} archivos reasignados ({rango_old}) a ({rango_new}) en la caja {caja_vis}"
        elif entidad == "caja" and accion == "ELIMINAR_CAJA":
            reasignados = despues.get("reasignados", 0)
            rango_old = f"{_fmt_num(antes.get('rango_min'))}-{_fmt_num(antes.get('rango_max'))}"
            texto = f"ELIMINAR_CAJA masivo: {reasignados} archivos reasignados desde la caja {caja_vis} ({rango_old})"
        elif entidad == "caja" and accion == "ELIMINAR_ARCHIVOS_CAJA":
            total_archivos = despues.get("archivos_eliminados", antes.get("total_archivos", 0))
            total_pdfs = despues.get("pdfs_eliminados", 0)
            texto = f"ELIMINAR_ARCHIVOS_CAJA permanente: {total_archivos} documentos y {total_pdfs} PDF(s) eliminados de la caja {caja_vis}"
        else:
            texto = f"{accion}"

        movimientos.append((fecha, texto, mid))

    cur.close()
    conn.close()

    return render_template(
        "archivos.html",
        archivos=datos,
        resultado=resultado,
        movimientos=movimientos,
        edit_num=edit_num
    )

# ---------------- ARCHIVO ----------------

@app.route("/archivo", methods=["GET", "POST"])
def archivo():
    if not login_requerido():
        return redirect(url_for("login"))

    grupo_id = obtener_grupo_id()
    if not grupo_id:
        return redirect(url_for("grupos"))

    asegurar_caja_sin_asignar(grupo_id)
    reparar_archivos_huerfanos(grupo_id)

    archivador_mode = admin_requerido() and es_archivador_grupo(grupo_id)
    view_mode = request.args.get("view", "").strip()
    import_job_id = request.args.get("import_job", "").strip() or session.get("last_import_job_id")
    import_job = (get_import_job(import_job_id) or get_import_report(import_job_id)) if import_job_id else None
    if import_job and not puede_ver_reporte_importacion(import_job):
        import_job = None
        import_job_id = None

    if request.method == "GET" and archivador_mode and view_mode == "especial":
        conn = get_db()
        cur = conn.cursor()

        cur.execute(
            """
            SELECT id, rango_min, rango_max, is_pendiente, grupo_origen_id
            FROM cajas
            WHERE grupo_id = %s
            ORDER BY is_pendiente, rango_min, id
            """,
            (grupo_id,)
        )
        cajas_rows = cur.fetchall()

        cur.execute(
            """
            SELECT id, numero, nombre, caja_id, pdf_path, grupo_origen_id
            FROM archivos
            WHERE grupo_id = %s
            ORDER BY caja_id, numero
            """,
            (grupo_id,)
        )
        archivos_rows = cur.fetchall()

        cur.execute(
            """
            SELECT id, nombre
            FROM grupos
            WHERE archivado = FALSE AND nombre <> 'Archivador'
            ORDER BY nombre
            """
        )
        grupos_destino = cur.fetchall()

        cur.execute("SELECT id, nombre FROM grupos")
        grupos_nombres = {r[0]: r[1] for r in cur.fetchall()}

        cur.close()
        conn.close()

        cajas_map = {}
        for c in cajas_rows:
            cajas_map[c[0]] = {
                "id": c[0],
                "rango_min": c[1],
                "rango_max": c[2],
                "is_pendiente": c[3],
                "grupo_origen_id": c[4],
                "grupo_origen_nombre": grupos_nombres.get(c[4]),
                "archivos": []
            }

        for a in archivos_rows:
            caja_id = a[3]
            if caja_id not in cajas_map:
                cajas_map[caja_id] = {
                    "id": caja_id,
                    "rango_min": None,
                    "rango_max": None,
                    "is_pendiente": False,
                    "grupo_origen_id": None,
                    "archivos": []
                }
            cajas_map[caja_id]["archivos"].append({
                "id": a[0],
                "numero": a[1],
                "nombre": a[2],
                "pdf_path": a[4],
                "grupo_origen_id": a[5],
                "grupo_origen_nombre": grupos_nombres.get(a[5]),
            })

        cajas_list = []
        for c in cajas_map.values():
            if c["is_pendiente"] and not c["archivos"]:
                continue
            cajas_list.append(c)

        return render_template(
            "archivo_archivador.html",
            cajas=cajas_list,
            grupos_destino=grupos_destino
        )

    # =========================================================
    # POST: acciones del dashboard
    # =========================================================
    accion = request.form.get("accion") if request.method == "POST" else None

    if request.method == "POST":
        # ---------- Importar Excel ----------
        if accion == "importar_excel":
            excel_file = request.files.get("excel")
            if not excel_file or not excel_file.filename:
                flash_error(500)
                return redirect(url_for("archivo"))

            uploads_dir = os.path.join(app.root_path, "uploads", "imports")
            os.makedirs(uploads_dir, exist_ok=True)
            ext = os.path.splitext(excel_file.filename)[1].lower()
            filename = f"import_{uuid.uuid4().hex}{ext or '.xlsx'}"
            file_path = os.path.join(uploads_dir, filename)
            excel_file.save(file_path)
            job_id = uuid.uuid4().hex
            group_name = get_group_name(grupo_id)
            session["last_import_job_id"] = job_id
            set_import_job(
                job_id,
                status="pending",
                message="Importacion iniciada.",
                source_filename=excel_file.filename,
                user_id=session.get("usuario_id"),
                user_name=session.get("usuario"),
                group_id=grupo_id,
                group_name=group_name,
                started_at=datetime.utcnow().isoformat(),
            )

            t = threading.Thread(
                target=importar_excel_job,
                args=(file_path, grupo_id, session.get("usuario_id"), job_id),
                daemon=True
            )
            t.start()

            flash("Importacion iniciada. Espera a que finalice la carga para continuar.", "info")
            return redirect(url_for("archivo", import_job=job_id))

        # ---------- Carga masiva PDF ----------
        if accion == "carga_masiva_pdf":
            pdf_files = request.files.getlist("pdfs")
            pdf_files = [f for f in pdf_files if f and f.filename]
            if not pdf_files:
                flash_error(422, fallback="Debes seleccionar al menos un PDF.")
                return redirect(url_for("archivo"))

            max_files = app.config["MASSIVE_PDF_MAX_FILES"]
            max_total_bytes = app.config["MASSIVE_PDF_MAX_TOTAL_BYTES"]
            total_bytes = sum(obtener_tamano_upload(file) for file in pdf_files)

            if len(pdf_files) > max_files:
                flash_error(
                    422,
                    fallback=f"Solo puedes subir hasta {max_files} PDFs por carga masiva."
                )
                return redirect(url_for("archivo"))

            if total_bytes > max_total_bytes:
                limite_mb = max_total_bytes / (1024 * 1024)
                total_mb = total_bytes / (1024 * 1024)
                flash_error(
                    422,
                    fallback=f"La carga masiva supera el limite de {limite_mb:.0f} MB. Tamaño actual: {total_mb:.1f} MB."
                )
                return redirect(url_for("archivo"))

            uploads_dir = os.path.join(app.root_path, "uploads", "imports_pdf", uuid.uuid4().hex)
            os.makedirs(uploads_dir, exist_ok=True)
            stored_names = []
            for index, file in enumerate(pdf_files, start=1):
                original_name = os.path.basename(file.filename or f"pdf_{index}.pdf")
                safe_name = secure_filename(original_name) or f"pdf_{index}.pdf"
                final_name = f"{index:04d}_{safe_name}"
                file.stream.seek(0)
                file.save(os.path.join(uploads_dir, final_name))
                stored_names.append(final_name)

            job_id = uuid.uuid4().hex
            group_name = get_group_name(grupo_id)
            session["last_import_job_id"] = job_id
            set_import_job(
                job_id,
                import_type="pdf",
                status="pending",
                message="Carga masiva de PDF iniciada.",
                source_filename=f"Carga masiva PDF ({len(pdf_files)} archivos)",
                user_id=session.get("usuario_id"),
                user_name=session.get("usuario"),
                group_id=grupo_id,
                group_name=group_name,
                total_rows=len(pdf_files),
                started_at=datetime.utcnow().isoformat(),
            )

            t = threading.Thread(
                target=importar_pdf_job,
                args=(uploads_dir, stored_names, grupo_id, session.get("usuario_id"), job_id, request.remote_addr),
                daemon=True
            )
            t.start()

            flash("Carga masiva de PDF iniciada. Espera a que finalice para continuar.", "info")
            return redirect(url_for("archivo", import_job=job_id))

        # ---------- Descarga masiva PDF ----------
        if accion == "descarga_masiva_pdf":
            selected_raw = request.form.get("selected_docs", "").strip()
            if not selected_raw:
                flash_error(422, fallback="Debes seleccionar al menos un PDF para descargar.")
                return redirect(url_for("archivo"))

            selected_docs = []
            for value in selected_raw.split(","):
                value = value.strip()
                if value.isdigit():
                    selected_docs.append(int(value))

            selected_docs = list(dict.fromkeys(selected_docs))
            if not selected_docs:
                flash_error(422)
                return redirect(url_for("archivo"))

            conn = get_db()
            cur = conn.cursor()
            in_clause, in_params = sqlserver_in_clause(selected_docs)
            cur.execute(
                f"""
                SELECT id, numero, nombre, pdf_path
                FROM archivos
                WHERE grupo_id = %s
                  AND id IN {in_clause}
                  AND pdf_path IS NOT NULL
                ORDER BY numero
                """,
                (grupo_id, *in_params)
            )
            rows = cur.fetchall()
            cur.close()
            conn.close()

            if not rows:
                flash_error(424)
                return redirect(url_for("archivo"))

            zip_buffer = BytesIO()
            agregados = 0
            with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
                for archivo_id, numero, nombre, pdf_path in rows:
                    abs_path = pdf_path
                    if not os.path.isabs(abs_path):
                        abs_path = os.path.join(app.config["UPLOAD_FOLDER"], abs_path)
                    if not os.path.exists(abs_path):
                        continue
                    safe_name = secure_filename(str(nombre or f"documento_{numero}")) or f"documento_{numero}"
                    zip_file.write(abs_path, arcname=f"{numero}_{safe_name}_{archivo_id}.pdf")
                    agregados += 1

            if not agregados:
                flash_error(425)
                return redirect(url_for("archivo"))

            zip_buffer.seek(0)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            return send_file(
                zip_buffer,
                as_attachment=True,
                download_name=f"pdfs_seleccionados_{timestamp}.zip",
                mimetype="application/zip"
            )

        # ---------- Crear Caja ----------
        if accion == "crear_caja":
            rmin = int(request.form["rango_min"])
            rmax = int(request.form["rango_max"])

            if not puede_crear_modificar_cajas():
                flash_error(205)
                return redirect(url_for("archivo"))

            nueva_caja_id = crear_caja(rmin, rmax, grupo_id, creado_por=session.get("usuario_id"))
            movidos = reubicar_archivos_pendientes_por_nueva_caja(
                nueva_caja_id,
                grupo_id,
                session.get("usuario_id")
            )

            registrar_log(
                session.get("usuario_id"),
                f"CREAR_CAJA caja_id={nueva_caja_id} rango={rmin}-{rmax} movidos_desde_caja0={movidos}",
                request.remote_addr,
                grupo_id
            )

            flash("Caja creada correctamente.", "success")
            return redirect(url_for("archivo"))

        # ---------- Crear Cajas Masivas ----------
        if accion == "crear_cajas_masivas":
            if not admin_requerido():
                flash_error(204)
                return redirect(url_for("archivo"))

            try:
                cantidad = int(request.form.get("cantidad_cajas", "0"))
                salto = int(request.form.get("salto_cajas", "0"))
                inicio = int(request.form.get("inicio_cajas", "0"))
            except (TypeError, ValueError):
                flash_error(400)
                return redirect(url_for("archivo"))

            if cantidad <= 0 or salto <= 0 or inicio < 0:
                flash_error(400)
                return redirect(url_for("archivo"))

            if cantidad > 500:
                flash_error(400, detail="La cantidad maxima por lote es 500 cajas.")
                return redirect(url_for("archivo"))

            creadas = []
            movidos_total = 0
            rango_inicio = inicio
            rango_fin = inicio + salto

            for _ in range(cantidad):
                caja_id = crear_caja(rango_inicio, rango_fin, grupo_id, creado_por=session.get("usuario_id"))
                movidos = reubicar_archivos_pendientes_por_nueva_caja(
                    caja_id,
                    grupo_id,
                    session.get("usuario_id")
                )
                creadas.append((caja_id, rango_inicio, rango_fin))
                movidos_total += movidos
                rango_inicio = rango_fin + 1
                rango_fin = rango_fin + salto

            primer_rango = f"{creadas[0][1]}-{creadas[0][2]}"
            ultimo_rango = f"{creadas[-1][1]}-{creadas[-1][2]}"

            registrar_log(
                session.get("usuario_id"),
                f"CREAR_CAJAS_MASIVAS cantidad={cantidad} salto={salto} inicio={inicio} movidos_desde_caja0={movidos_total} primer_rango={primer_rango} ultimo_rango={ultimo_rango}",
                request.remote_addr,
                grupo_id
            )

            flash(
                f"Se crearon {cantidad} cajas correctamente. Rango inicial {primer_rango}, rango final {ultimo_rango}.",
                "success"
            )
            return redirect(url_for("archivo"))

        # ---------- Agregar Documento ----------
        elif accion == "agregar_documento":
            numero = int(request.form["numero"])
            tipo_doc = normalizar_tipo_doc(request.form.get("tipo_doc", ""))
            if not es_tipo_doc_valido(tipo_doc):
                flash_error(400)
                return redirect(url_for("archivo"))
            nombre = request.form.get("nombre", "").strip()
            if not nombre:
                nombre = f"Documento {numero}"
            nombre = normalizar_nombre(nombre)

            conn = get_db()
            cur = conn.cursor()

            # 1) Determinar caja destino por rangos
            cur.execute("""
                SELECT id
                FROM cajas
                WHERE grupo_id = %s
                  AND is_pendiente = FALSE
                  AND %s BETWEEN rango_min AND rango_max
                ORDER BY rango_min, id
                LIMIT 1
            """, (grupo_id, numero))
            caja_dest = cur.fetchone()
            caja_id = caja_dest[0] if caja_dest else asegurar_caja_sin_asignar(grupo_id)  # si no cae en ninguna caja → 0

            # 2) Insertar archivo YA con caja_id
            cur.execute("""
                INSERT INTO archivos (caja_id, numero, nombre, pdf_path, grupo_id, creado_por, tipo_doc)
                VALUES (%s, %s, %s, NULL, %s, %s, %s)
                RETURNING id
            """, (caja_id, numero, nombre, grupo_id, session.get("usuario_id"), tipo_doc))
            archivo_id = cur.fetchone()[0]

            # 3) PDF opcional: guardar y actualizar pdf_path
            pdf_name = None
            file = request.files.get("pdf")
            if file and file.filename:
                if not es_pdf(file):
                    cur.close()
                    conn.close()
                    flash_error(407)
                    return redirect(url_for("archivo"))

                pdf_name = guardar_pdf(file, numero, grupo_id)
                cur.execute(
                    "UPDATE archivos SET pdf_path = %s WHERE numero = %s AND grupo_id = %s",
                    (pdf_name, numero, grupo_id)
                )

            conn.commit()
            cur.close()
            conn.close()

            registrar_log(
                session.get("usuario_id"),
                f"ARCHIVO|tipo=CREACION|numero={numero}|nombre={nombre}|caja={caja_id}",
                request.remote_addr,
                grupo_id
            )

            registrar_movimiento(
                session.get("usuario_id"),
                grupo_id,
                entidad="archivo",
                entidad_id=archivo_id,
                accion="CREAR_ARCHIVO",
                datos_despues={
                    "id": archivo_id,
                    "numero": numero,
                    "nombre": nombre,
                    "caja_id": caja_id,
                    "pdf_path": pdf_name,
                    "tipo_doc": tipo_doc,
                },
            )

            flash("Documento agregado correctamente.", "success")
            return redirect(url_for("archivo"))

        # ---------- Eliminar Archivo (desde modal de bÃºsqueda) ----------
        elif accion == "eliminar_archivo_modal":
            numero = int(request.form["numero"])

            if not supervisor_requerido() and not usuario_puede_eliminar(session.get("usuario_id"), grupo_id):
                flash_error(203)
                return redirect(url_for("archivo"))

            conn = get_db()
            cur = conn.cursor()

            # datos antes de borrar para log + pdf_path para borrar del disco
            cur.execute(
                "SELECT id, caja_id, numero, nombre, pdf_path, tipo_doc FROM archivos WHERE numero = %s AND grupo_id = %s",
                (numero, grupo_id)
            )
            antes = cur.fetchone()

            if antes:
                archivo_id, caja_old, numero_old, nombre_old, pdf_old, tipo_doc_old = antes
                registrar_movimiento(
                    session.get("usuario_id"),
                    grupo_id,
                    entidad="archivo",
                    entidad_id=archivo_id,
                    accion="ELIMINAR_ARCHIVO",
                    datos_antes={
                        "id": archivo_id,
                        "numero": numero_old,
                        "nombre": nombre_old,
                        "caja_id": caja_old,
                        "pdf_path": pdf_old,
                        "tipo_doc": tipo_doc_old,
                    },
                )

            cur.execute("DELETE FROM archivos WHERE numero = %s AND grupo_id = %s", (numero, grupo_id))
            conn.commit()
            cur.close()
            conn.close()

            # borrar PDF del disco si existÃ­a
            if antes:

                if pdf_old:
                    try:
                        path = pdf_old
                        if not os.path.isabs(path):
                            path = os.path.join(app.root_path, path)
                        if os.path.exists(path):
                            os.remove(path)
                    except Exception as e:
                        print("Error eliminando PDF:", e)

                registrar_log(
                    session.get("usuario_id"),
                    f"ARCHIVO|tipo=ELIMINACION|numero_old={numero_old}|nombre_old={nombre_old}|caja={caja_old}",
                    request.remote_addr,
                    grupo_id
                )

            flash("Archivo eliminado correctamente.", "success")
            return redirect(url_for("archivo"))

        # ---------- Modificar Archivo (desde modal de bÃºsqueda) ----------
        elif accion == "modificar_archivo_modal":
            numero_old = int(request.form["numero_old"])
            numero_new = int(request.form["numero_new"])
            nombre_new = request.form.get("nombre_new", "").strip()
            nombre_new = normalizar_nombre(nombre_new)
            tipo_doc_new = normalizar_tipo_doc(request.form.get("tipo_doc_new", ""))
            if not es_tipo_doc_valido(tipo_doc_new):
                flash_error(400)
                return redirect(url_for("archivo"))

            # checkbox para eliminar PDF actual / agregar PDF
            remove_pdf = request.form.get("remove_pdf") == "1"
            append_pdf = request.form.get("append_pdf") == "1"
            if append_pdf:
                remove_pdf = False

            conn = get_db()
            cur = conn.cursor()

            cur.execute(
                "SELECT id, caja_id, numero, nombre, pdf_path, tipo_doc FROM archivos WHERE numero = %s AND grupo_id = %s",
                (numero_old, grupo_id)
            )
            antes = cur.fetchone()

            if not antes:
                cur.close()
                conn.close()
                flash_error(404)
                return redirect(url_for("archivo"))

            archivo_id, caja_old, numero_viejo, nombre_viejo, pdf_old, tipo_doc_old = antes

            # Caja destino segÃºn rangos del numero_new
            cur.execute("""
                SELECT id
                FROM cajas
                WHERE grupo_id = %s
                  AND is_pendiente = FALSE
                  AND %s BETWEEN rango_min AND rango_max
                ORDER BY rango_min, id
                LIMIT 1
            """, (grupo_id, numero_new))
            caja_dest = cur.fetchone()
            caja_dest_id = caja_dest[0] if caja_dest else asegurar_caja_sin_asignar(grupo_id)

            pdf_name = pdf_old

            # 1) eliminar PDF actual si se pidiÃ³
            if remove_pdf and pdf_old:
                try:
                    path = pdf_old
                    if not os.path.isabs(path):
                        path = os.path.join(app.root_path, path)
                    if os.path.exists(path):
                        os.remove(path)
                except Exception as e:
                    print("Error eliminando PDF:", e)

                pdf_name = None  # queda NULL en DB

            # 2) subir PDF nuevo (reemplazo o agregar)
            file = request.files.get("pdf")
            if file and file.filename:
                if not es_pdf(file):
                    cur.close()
                    conn.close()
                    flash_error(407)
                    return redirect(url_for("archivo"))

                if append_pdf:
                    pdf_name = unir_pdf_existente(pdf_old, file, numero_new, grupo_id)
                    if not pdf_name:
                        cur.close()
                        conn.close()
                        flash_error(423)
                        return redirect(url_for("archivo"))
                else:
                    # si habÃ­a anterior y no se eliminÃ³ arriba, lo borramos
                    if pdf_old and not remove_pdf:
                        try:
                            old_path = pdf_old
                            if not os.path.isabs(old_path):
                                old_path = os.path.join(app.root_path, old_path)
                            if os.path.exists(old_path):
                                os.remove(old_path)
                        except Exception as e:
                            print("Error eliminando PDF anterior:", e)

                    pdf_name = guardar_pdf(file, numero_new, grupo_id)

            # 3) update final
            cur.execute("""
                UPDATE archivos
                SET numero = %s, nombre = %s, caja_id = %s, pdf_path = %s, tipo_doc = %s
                WHERE numero = %s AND grupo_id = %s
            """, (numero_new, nombre_new, caja_dest_id, pdf_name, tipo_doc_new, numero_old, grupo_id))

            conn.commit()
            cur.close()
            conn.close()

            registrar_log(
                session.get("usuario_id"),
                f"ARCHIVO|tipo=MODIFICACION|numero_old={numero_viejo}|numero_new={numero_new}|"
                f"nombre_old={nombre_viejo}|nombre_new={nombre_new}|caja={caja_dest_id}|"
                f"pdf_eliminado={1 if (remove_pdf and pdf_old) else 0}|pdf_nuevo={1 if (file and file.filename) else 0}|pdf_unido={1 if append_pdf else 0}",
                request.remote_addr,
                grupo_id
            )

            registrar_movimiento(
                session.get("usuario_id"),
                grupo_id,
                entidad="archivo",
                entidad_id=archivo_id,
                accion="MODIFICAR_ARCHIVO",
                datos_antes={
                    "numero": numero_viejo,
                    "nombre": nombre_viejo,
                    "caja_id": caja_old,
                    "pdf_path": pdf_old,
                    "tipo_doc": tipo_doc_old,
                },
                datos_despues={
                    "numero": numero_new,
                    "nombre": nombre_new,
                    "caja_id": caja_dest_id,
                    "pdf_path": pdf_name,
                    "tipo_doc": tipo_doc_new,
                },
            )

            flash("Archivo modificado correctamente.", "success")
            return redirect(url_for("archivo"))

        # AcciÃ³n desconocida
        else:
            flash_error(200)
            return redirect(url_for("archivo"))

    # =========================================================
    # GET: buscador + listado de cajas
    # =========================================================
    resultado = None
    buscar_raw = request.args.get("buscar", "").strip()

    if buscar_raw:
        conn = get_db()
        cur = conn.cursor()
        try:
            buscar_clean = buscar_raw.replace(".", "")
            if buscar_clean.isdigit():
                doc = int(buscar_clean)
                if "." in buscar_raw:
                    cur.execute("""
                        WITH ranked AS (
                            SELECT id, ROW_NUMBER() OVER (ORDER BY rango_min, id) AS caja_visible
                            FROM cajas
                            WHERE grupo_id = %s AND is_pendiente = FALSE
                        )
                    SELECT
                    c.id AS caja_id,
                    CASE WHEN c.is_pendiente = 1 THEN 0 ELSE r.caja_visible END AS caja_num,
                    a.tipo_doc AS tipo_doc,
                    a.numero AS documento,
                    a.nombre AS nombre,
                    a.pdf_path
                    FROM archivos a
                        JOIN cajas c ON c.id = a.caja_id
                        LEFT JOIN ranked r ON r.id = c.id
                        WHERE a.grupo_id = %s
                          AND (
                            a.numero = %s
                            OR a.nombre ILIKE %s
                          )
                        ORDER BY a.numero
                    """, (grupo_id, grupo_id, doc, f"%{buscar_raw}%"))
                else:
                    cur.execute("""
                        WITH ranked AS (
                            SELECT id, ROW_NUMBER() OVER (ORDER BY rango_min, id) AS caja_visible
                            FROM cajas
                            WHERE grupo_id = %s AND is_pendiente = FALSE
                        )
                    SELECT
                    c.id AS caja_id,
                    CASE WHEN c.is_pendiente = 1 THEN 0 ELSE r.caja_visible END AS caja_num,
                    a.tipo_doc AS tipo_doc,
                    a.numero AS documento,
                    a.nombre AS nombre,
                    a.pdf_path
                    FROM archivos a
                        JOIN cajas c ON c.id = a.caja_id
                        LEFT JOIN ranked r ON r.id = c.id
                        WHERE a.grupo_id = %s
                          AND (
                            a.numero = %s
                            OR replace(a.nombre, '.', '') = %s
                          )
                        ORDER BY a.numero
                    """, (grupo_id, grupo_id, doc, buscar_clean))

                rows = cur.fetchall()
                if rows:
                    resultado = rows_to_builtin(rows)
                else:
                    resultado = ("no",)
            else:
                cur.execute("""
                    WITH ranked AS (
                        SELECT id, ROW_NUMBER() OVER (ORDER BY rango_min, id) AS caja_visible
                        FROM cajas
                        WHERE grupo_id = %s AND is_pendiente = FALSE
                    )
                SELECT
                c.id AS caja_id,
                CASE WHEN c.is_pendiente = 1 THEN 0 ELSE r.caja_visible END AS caja_num,
                a.tipo_doc AS tipo_doc,
                a.numero AS documento,
                a.nombre AS nombre,
                a.pdf_path
                FROM archivos a
                    JOIN cajas c ON c.id = a.caja_id
                    LEFT JOIN ranked r ON r.id = c.id
                    WHERE a.nombre ILIKE %s AND a.grupo_id = %s
                    ORDER BY a.numero
                """, (grupo_id, f"%{buscar_raw}%", grupo_id))
                rows = cur.fetchall()
                if rows:
                    resultado = rows_to_builtin(rows)
                else:
                    resultado = ("no",)
        finally:
            cur.close()
            conn.close()

    # listado de cajas (como ya lo tienes)
    conn = get_db()
    cur = conn.cursor()
    cur.execute("""
        WITH ranked AS (
            SELECT id, ROW_NUMBER() OVER (ORDER BY rango_min, id) AS caja_visible
            FROM cajas
            WHERE grupo_id = %s AND is_pendiente = FALSE
        ),
        conteo AS (
            SELECT caja_id, COUNT(*) AS total_archivos
            FROM archivos
            WHERE grupo_id = %s
            GROUP BY caja_id
        )
        SELECT
            c.id,
            CASE
            WHEN c.is_pendiente = 1 THEN 0
            ELSE r.caja_visible
            END AS caja_num,
            c.rango_min,
            c.rango_max,
            COALESCE(ct.total_archivos, 0) AS total_archivos,
            c.is_pendiente
        FROM cajas c
        LEFT JOIN conteo ct ON ct.caja_id = c.id
        LEFT JOIN ranked r ON r.id = c.id
        WHERE c.grupo_id = %s AND (c.is_pendiente = FALSE OR COALESCE(ct.total_archivos, 0) > 0)
        ORDER BY CASE WHEN c.is_pendiente = 1 THEN 1 ELSE 0 END, c.rango_min, c.id
    """, (grupo_id, grupo_id, grupo_id))

    cajas = cur.fetchall()

    cajas_map = {c[0]: c[1] for c in cajas}

    cur.execute(
        """
        SELECT a.id, a.caja_id, a.numero, a.nombre, a.tipo_doc
        FROM archivos a
        WHERE a.grupo_id = %s
          AND a.pdf_path IS NOT NULL
        ORDER BY a.caja_id, a.numero, a.id
        """,
        (grupo_id,)
    )
    pdf_rows = cur.fetchall()
    cur.close()
    conn.close()

    pdf_bulk_data = []
    docs_by_box = {}
    for archivo_id, caja_id, numero, nombre, tipo_doc in pdf_rows:
        docs_by_box.setdefault(caja_id, []).append({
            "id": archivo_id,
            "numero": numero,
            "nombre": nombre,
            "tipo_doc": tipo_doc,
        })

    for caja in cajas:
        caja_id = caja[0]
        docs = docs_by_box.get(caja_id, [])
        if not docs:
            continue
        pdf_bulk_data.append({
            "id": caja_id,
            "numero": caja[1],
            "docs": docs,
        })

    return render_template(
        "archivo_dashboard.html",
        cajas=cajas,
        resultado=resultado,
        archivador_mode=archivador_mode,
        pdf_bulk_data=pdf_bulk_data,
        import_job=import_job,
        import_status_url=url_for("importacion_estado", job_id=import_job_id) if import_job_id else None,
        import_report_url_base=url_for("importacion_reporte", job_id="__JOB__"),
    )

# ---------------- archivo_caja ----------------

@app.route("/archivo/caja/<int:caja_id>", methods=["GET", "POST"])
def archivo_caja(caja_id):
    if not login_requerido():
        return redirect(url_for("login"))

    grupo_id = obtener_grupo_id()
    if not grupo_id:
        return redirect(url_for("grupos"))

    asegurar_caja_sin_asignar(grupo_id)
    reparar_archivos_huerfanos(grupo_id)

    # ======================
    # POST: Acciones en esta caja
    # ======================
    if request.method == "POST":
        accion = request.form.get("accion")

        # ---------- MODIFICAR CAJA ----------
        if accion == "modificar_caja":
            nuevo_min = int(request.form["rango_min"])
            nuevo_max = int(request.form["rango_max"])

            if not puede_crear_modificar_cajas():
                flash_error(204)
                return redirect(url_for("archivo_caja", caja_id=caja_id))

            caja_antes = modificar_caja(caja_id, nuevo_min, nuevo_max, grupo_id, session.get("usuario_id"))

            # Reubicar archivos segÃºn nuevo rango
            movidos_fuera = reubicar_archivos_de_caja(caja_id, grupo_id, session.get("usuario_id"))
            movidos_dentro = reubicar_archivos_pendientes_por_nueva_caja(
                caja_id,
                grupo_id,
                session.get("usuario_id")
            )
            total = movidos_fuera + movidos_dentro

            if caja_antes:
                registrar_movimiento(
                    session.get("usuario_id"),
                    grupo_id,
                    entidad="caja",
                    entidad_id=caja_id,
                    accion="MODIFICAR_CAJA",
                    datos_antes={
                        "rango_min": caja_antes["rango_min"],
                        "rango_max": caja_antes["rango_max"],
                    },
                    datos_despues={
                        "rango_min": caja_antes["nuevo_min"],
                        "rango_max": caja_antes["nuevo_max"],
                        "reasignados": total,
                        "modo": "masivo",
                    },
                )

            registrar_log(
                session.get("usuario_id"),
                f"MODIFICAR_CAJA_MASIVO caja={caja_id} movidos={total}",
                request.remote_addr,
                grupo_id
            )

            if total > 0:
                flash(f"Se reasignaron automaticamente {total} archivo(s) segun el nuevo rango.", "success")
            else:
                flash("El rango se actualizo. No fue necesario reasignar archivos.", "info")

            return redirect(url_for("archivo_caja", caja_id=caja_id))

        # ---------- ELIMINAR CAJA ----------
        if accion == "eliminar_caja":
            if not puede_eliminar_cajas():
                flash_error(355, fallback="Solo el supervisor o admin puede eliminar cajas.")
                return redirect(url_for("archivo_caja", caja_id=caja_id))

            caja_eliminada = eliminar_caja(caja_id, grupo_id, session.get("usuario_id"))

            if caja_eliminada:
                registrar_movimiento(
                    session.get("usuario_id"),
                    grupo_id,
                    entidad="caja",
                    entidad_id=caja_id,
                    accion="ELIMINAR_CAJA",
                    datos_antes={
                        "rango_min": caja_eliminada["rango_min"],
                        "rango_max": caja_eliminada["rango_max"],
                    },
                    datos_despues={
                        "reasignados": caja_eliminada["reasignados"],
                        "modo": "masivo",
                    },
                )

            registrar_log(
                session.get("usuario_id"),
                f"ELIMINAR_CAJA_MASIVO caja={caja_id} movidos={caja_eliminada['reasignados'] if caja_eliminada else 0}",
                request.remote_addr,
                grupo_id
            )

            flash("Caja eliminada correctamente.", "success")
            return redirect(url_for("archivo"))

        # ---------- ELIMINAR TODOS LOS ARCHIVOS DE LA CAJA ----------
        if accion == "eliminar_todos_archivos_caja":
            if not admin_requerido():
                flash_error(200, fallback="Solo el admin puede eliminar todos los archivos de una caja.")
                return redirect(url_for("archivo_caja", caja_id=caja_id))

            admin_delete_password = request.form.get("admin_delete_password", "")
            if admin_delete_password != app.config["ADMIN_BULK_DELETE_PASSWORD"]:
                flash_error(200, fallback="Clave de seguridad invalida. No se eliminaron los documentos.")
                return redirect(url_for("archivo_caja", caja_id=caja_id))

            conn = get_db()
            cur = conn.cursor()
            cur.execute(
                """
                SELECT id, numero, pdf_path
                FROM archivos
                WHERE caja_id = %s AND grupo_id = %s
                """,
                (caja_id, grupo_id)
            )
            archivos_caja = cur.fetchall()

            total_archivos = len(archivos_caja)
            pdfs = []
            for _, _, pdf_path in archivos_caja:
                if pdf_path:
                    abs_path = pdf_path
                    if not os.path.isabs(abs_path):
                        abs_path = os.path.join(app.config["UPLOAD_FOLDER"], pdf_path)
                    pdfs.append(abs_path)

            cur.execute(
                "DELETE FROM archivos WHERE caja_id = %s AND grupo_id = %s",
                (caja_id, grupo_id)
            )
            conn.commit()
            cur.close()
            conn.close()

            eliminados_pdf = 0
            for path in pdfs:
                try:
                    if os.path.exists(path):
                        os.remove(path)
                        eliminados_pdf += 1
                except Exception:
                    app.logger.exception("No se pudo eliminar PDF de borrado masivo")

            registrar_movimiento(
                session.get("usuario_id"),
                grupo_id,
                entidad="caja",
                entidad_id=caja_id,
                accion="ELIMINAR_ARCHIVOS_CAJA",
                datos_antes={
                    "total_archivos": total_archivos,
                    "total_pdfs": len(pdfs),
                },
                datos_despues={
                    "archivos_eliminados": total_archivos,
                    "pdfs_eliminados": eliminados_pdf,
                    "modo": "permanente",
                },
            )

            registrar_log(
                session.get("usuario_id"),
                f"ELIMINAR_ARCHIVOS_CAJA caja={caja_id} archivos={total_archivos} pdfs={eliminados_pdf}",
                request.remote_addr,
                grupo_id
            )

            if total_archivos == 0:
                flash("La caja ya no tenia documentos.", "info")
            else:
                flash(f"Se eliminaron permanentemente {total_archivos} documento(s) de la caja.", "success")
            return redirect(url_for("archivo_caja", caja_id=caja_id))

        # ---------- ELIMINAR ARCHIVO ----------
        if accion == "eliminar_archivo_fila":
            numero = int(request.form["numero"])

            if not supervisor_requerido() and not usuario_puede_eliminar(session.get("usuario_id"), grupo_id):
                flash_error(203)
                return redirect(url_for("archivo_caja", caja_id=caja_id))

            conn = get_db()
            cur = conn.cursor()

            # obtener datos antes de borrar (para log)
            cur.execute(
                "SELECT id, caja_id, numero, nombre, pdf_path, tipo_doc FROM archivos WHERE numero = %s AND grupo_id = %s",
                (numero, grupo_id)
            )
            antes = cur.fetchone()

            if antes:
                archivo_id, caja_old, numero_old, nombre_old, pdf_old, tipo_doc_old = antes
                registrar_movimiento(
                    session.get("usuario_id"),
                    grupo_id,
                    entidad="archivo",
                    entidad_id=archivo_id,
                    accion="ELIMINAR_ARCHIVO",
                    datos_antes={
                        "id": archivo_id,
                        "numero": numero_old,
                        "nombre": nombre_old,
                        "caja_id": caja_old,
                        "pdf_path": pdf_old,
                        "tipo_doc": tipo_doc_old,
                    },
                )

            cur.execute("DELETE FROM archivos WHERE numero = %s AND grupo_id = %s", (numero, grupo_id))
            conn.commit()

            cur.close()
            conn.close()

            if antes:
                registrar_log(
                    session.get("usuario_id"),
                    f"ARCHIVO|tipo=ELIMINACION|numero_old={numero_old}|nombre_old={nombre_old}|caja={caja_old}",
                    request.remote_addr,
                    grupo_id
                )

            flash("Archivo eliminado correctamente.", "success")
            return redirect(url_for("archivo_caja", caja_id=caja_id))

        # ---------- MODIFICAR ARCHIVO (numero y/o nombre) ----------
        if accion == "modificar_archivo_fila":
            import os

            numero_old = int(request.form["numero_old"])
            numero_new = int(request.form["numero_new"])
            nombre_new = request.form.get("nombre_new", "").strip()
            tipo_doc_new = normalizar_tipo_doc(request.form.get("tipo_doc_new", ""))
            if not es_tipo_doc_valido(tipo_doc_new):
                flash_error(400)
                return redirect(url_for("archivo_caja", caja_id=caja_id))

            remove_pdf = request.form.get("remove_pdf") == "1"
            append_pdf = request.form.get("append_pdf") == "1"
            if append_pdf:
                remove_pdf = False
            file = request.files.get("pdf")

            conn = get_db()
            cur = conn.cursor()

            # Traer estado actual
            cur.execute(
                "SELECT id, caja_id, numero, nombre, pdf_path, tipo_doc FROM archivos WHERE numero = %s AND grupo_id = %s",
                (numero_old, grupo_id)
            )
            row = cur.fetchone()
            if not row:
                cur.close()
                conn.close()
                flash_error(404)
                return redirect(url_for("archivo_caja", caja_id=caja_id))

            archivo_id, caja_old, numero_viejo, nombre_viejo, pdf_old, tipo_doc_old = row

            pdf_name = pdf_old

            # ðŸ—‘ï¸ Eliminar PDF actual
            if remove_pdf and pdf_old:
                try:
                    path = pdf_old
                    if not os.path.isabs(path):
                        path = os.path.join(app.root_path, path)
                    if os.path.exists(path):
                        os.remove(path)
                except Exception as e:
                    print("Error eliminando PDF:", e)

                pdf_name = None

            # ðŸ“„ Reemplazar / agregar PDF
            if file and file.filename:
                if not es_pdf(file):
                    cur.close()
                    conn.close()
                    flash_error(407)
                    return redirect(url_for("archivo_caja", caja_id=caja_id))

                if append_pdf:
                    pdf_name = unir_pdf_existente(pdf_old, file, numero_new, grupo_id)
                    if not pdf_name:
                        cur.close()
                        conn.close()
                        flash_error(423)
                        return redirect(url_for("archivo_caja", caja_id=caja_id))
                else:
                    # borrar anterior si existÃ­a
                    if pdf_old and not remove_pdf:
                        try:
                            old_path = pdf_old
                            if not os.path.isabs(old_path):
                                old_path = os.path.join(app.root_path, old_path)
                            if os.path.exists(old_path):
                                os.remove(old_path)
                        except Exception as e:
                            print("Error eliminando PDF anterior:", e)

                    pdf_name = guardar_pdf(file, numero_new, grupo_id)

            # Update final
            cur.execute("""
                UPDATE archivos
                SET numero = %s,
                    nombre = %s,
                    pdf_path = %s,
                    tipo_doc = %s
                WHERE numero = %s AND grupo_id = %s
            """, (numero_new, nombre_new, pdf_name, tipo_doc_new, numero_old, grupo_id))

            conn.commit()
            cur.close()
            conn.close()
            
            registrar_movimiento(
                session.get("usuario_id"),
                grupo_id,
                entidad="archivo",
                entidad_id=archivo_id,
                accion="MODIFICAR_ARCHIVO",
                datos_antes={
                    "numero": numero_viejo,
                    "nombre": nombre_viejo,
                    "caja_id": caja_old,
                    "pdf_path": pdf_old,
                    "tipo_doc": tipo_doc_old,
                },
                datos_despues={
                    "numero": numero_new,
                    "nombre": nombre_new,
                    "caja_id": caja_old,
                    "pdf_path": pdf_name,
                    "tipo_doc": tipo_doc_new,
                },
            )

            flash("Archivo modificado correctamente.", "success")
            return redirect(url_for("archivo_caja", caja_id=caja_id))

    # ======================
    # GET: Render de la caja + archivos
    # ======================
    highlight_raw = request.args.get("highlight", "").strip()
    highlight_num = int(highlight_raw) if highlight_raw.isdigit() else None

    conn = get_db()
    cur = conn.cursor()

    # info de caja + numero visible
    cur.execute("""
        WITH ranked AS (
            SELECT id, ROW_NUMBER() OVER (ORDER BY rango_min, id) AS caja_visible
            FROM cajas
            WHERE grupo_id = %s AND is_pendiente = FALSE
        )
        SELECT
            c.id,
            CASE WHEN c.is_pendiente = 1 THEN 0 ELSE r.caja_visible END AS caja_num,
            c.rango_min,
            c.rango_max
        FROM cajas c
        LEFT JOIN ranked r ON r.id = c.id
        WHERE c.id = %s AND c.grupo_id = %s
    """, (grupo_id, caja_id, grupo_id))
    caja_info = cur.fetchone()

    if not caja_info:
        cur.close()
        conn.close()
        flash_error(350)
        return redirect(url_for("archivo"))

    # archivos de la caja (incluye pdf_path para habilitar Ver PDF)
    cur.execute("""
        SELECT a.numero, a.tipo_doc, a.nombre, a.pdf_path
        FROM archivos a
        WHERE a.caja_id = %s AND a.grupo_id = %s
        ORDER BY a.numero
    """, (caja_id, grupo_id))
    archivos = cur.fetchall()

    cur.close()
    conn.close()

    return render_template(
        "archivo_caja.html",
        caja=caja_info,
        archivos=archivos,
        highlight_num=highlight_num
    )


# ---------------- ARCHIVO: DUPLICADOS ----------------
@app.route("/archivo/duplicados", methods=["GET", "POST"])
def archivo_duplicados():
    if not login_requerido():
        return redirect(url_for("login"))

    if not admin_requerido():
        return "Acceso denegado", 403

    grupo_id = obtener_grupo_id()
    if not grupo_id:
        return redirect(url_for("grupos"))

    if request.method == "POST":
        accion = request.form.get("accion")
        if accion == "unificar":
            ids_raw = request.form.get("ids", "")
            base_id_raw = request.form.get("base_id", "")
            tipo_doc = normalizar_tipo_doc(request.form.get("tipo_doc", ""))
            numero_raw = request.form.get("numero", "").strip()
            nombre = request.form.get("nombre", "").strip()

            if not ids_raw:
                flash_error(550)
                return redirect(url_for("archivo_duplicados"))

            ids = [int(x) for x in ids_raw.split(",") if x.strip().isdigit()]
            if len(ids) < 2:
                flash_error(551)
                return redirect(url_for("archivo_duplicados"))

            if not base_id_raw or not base_id_raw.isdigit():
                flash_error(552)
                return redirect(url_for("archivo_duplicados"))

            base_id = int(base_id_raw)
            if base_id not in ids:
                flash_error(553)
                return redirect(url_for("archivo_duplicados"))

            if not es_tipo_doc_valido(tipo_doc):
                flash_error(400)
                return redirect(url_for("archivo_duplicados"))

            if not numero_raw.isdigit():
                flash_error(401)
                return redirect(url_for("archivo_duplicados"))

            if not nombre:
                flash_error(402)
                return redirect(url_for("archivo_duplicados"))

            numero = int(numero_raw)
            nombre = normalizar_nombre(nombre)

            conn = get_db()
            cur = conn.cursor()
            try:
                rows = []
                for chunk in iter_chunks(ids):
                    in_clause, in_params = sqlserver_in_clause(chunk)
                    cur.execute(
                        f"SELECT id FROM archivos WHERE id IN {in_clause} AND grupo_id = %s",
                        (*in_params, grupo_id)
                    )
                    rows.extend([r[0] for r in cur.fetchall()])
                if len(set(rows)) != len(set(ids)):
                    flash_error(554)
                    return redirect(url_for("archivo_duplicados"))

                otros = [i for i in ids if i != base_id]
                if otros:
                    for chunk in iter_chunks(otros):
                        in_clause, in_params = sqlserver_in_clause(chunk)
                        cur.execute(
                            f"DELETE FROM archivos WHERE id IN {in_clause} AND grupo_id = %s",
                            (*in_params, grupo_id)
                        )

                cur.execute(
                    """
                    UPDATE archivos
                    SET tipo_doc = %s, numero = %s, nombre = %s
                    WHERE id = %s AND grupo_id = %s
                    """,
                    (tipo_doc, numero, nombre, base_id, grupo_id)
                )

                conn.commit()
                flash("Registros unificados correctamente.", "success")
            except Exception:
                conn.rollback()
                app.logger.exception("Error unificando duplicados")
                flash_error(555)
            finally:
                cur.close()
                conn.close()

            return redirect(url_for("archivo_duplicados"))

    # ======= GET: detectar duplicados =======
    conn = get_db()
    cur = conn.cursor()
    cur.execute(
        """
        SELECT id, tipo_doc, numero, nombre
        FROM archivos
        WHERE grupo_id = %s
        ORDER BY numero, id
        """,
        (grupo_id,)
    )
    rows = cur.fetchall()
    cur.close()
    conn.close()

    items = [
        {
            "id": r[0],
            "tipo_doc": r[1],
            "numero": r[2],
            "nombre": r[3] or ""
        }
        for r in rows
    ]

    grupos = []
    usados = set()

    # --- 1) Duplicados por numero (sin importar tipo_doc)
    por_numero = defaultdict(list)
    for it in items:
        por_numero[it["numero"]].append(it)
    for numero, lst in por_numero.items():
        if len(lst) > 1:
            grupos.append({
                "tipo": "numero",
                "clave": str(numero),
                "items": lst
            })
            usados.update([x["id"] for x in lst])

    # --- 2) Duplicados por nombre similar
    candidatos = [it for it in items if it["id"] not in usados and it["nombre"]]
    # bucket por prefijo para limitar comparaciones
    buckets = defaultdict(list)
    for it in candidatos:
        key = normalizar_nombre_clave(it["nombre"]).replace(" ", "")
        pref = key[:3] if len(key) >= 3 else key
        buckets[pref].append(it)

    # Union-Find simple
    parent = {it["id"]: it["id"] for it in candidatos}

    def find(x):
        while parent[x] != x:
            parent[x] = parent[parent[x]]
            x = parent[x]
        return x

    def union(a, b):
        ra, rb = find(a), find(b)
        if ra != rb:
            parent[rb] = ra

    def doc_sim(a, b):
        sa = str(a) if a is not None else ""
        sb = str(b) if b is not None else ""
        if not sa or not sb:
            return 0.0
        return SequenceMatcher(None, sa, sb).ratio()

    for pref, lst in buckets.items():
        n = len(lst)
        for i in range(n):
            name_i = normalizar_nombre_clave(lst[i]["nombre"])
            for j in range(i + 1, n):
                name_j = normalizar_nombre_clave(lst[j]["nombre"])
                if abs(len(name_i) - len(name_j)) > 4:
                    continue
                ratio = SequenceMatcher(None, name_i, name_j).ratio()
                if ratio >= 0.88 and doc_sim(lst[i]["numero"], lst[j]["numero"]) >= 0.75:
                    union(lst[i]["id"], lst[j]["id"])

    grupos_nombre = defaultdict(list)
    for it in candidatos:
        grupos_nombre[find(it["id"])].append(it)

    for _, lst in grupos_nombre.items():
        if len(lst) > 1:
            grupos.append({
                "tipo": "nombre",
                "clave": normalizar_nombre_clave(lst[0]["nombre"]),
                "items": lst
            })

    return render_template(
        "archivo_duplicados.html",
        grupos=grupos,
        tipo_doc_opciones=TIPO_DOC_OPCIONES
    )


# ---------------- ADMIN: LOGS ----------------
@app.route("/admin/logs")
def admin_logs():
    if not login_requerido():
        return redirect(url_for("login"))

    if not admin_requerido():
        return "Acceso denegado", 403

    conn = get_db()
    cur = conn.cursor()

    registros = []
    movimientos = []
    usuarios = []
    grupos_todos = []
    logs_por_grupo = defaultdict(list)
    movimientos_por_grupo = defaultdict(list)

    try:
        cur.execute("""
            SELECT
                COALESCE(u.usuario, '[eliminado]') AS usuario,
                a.accion,
                a.fecha,
                a.direccionip,
                e.nombre,
                a.empresa
            FROM auditoria a
            LEFT JOIN usuarios u ON u.id = a.usuarioid
            LEFT JOIN empresas e ON e.id = a.empresa
            ORDER BY a.fecha DESC
            LIMIT 200
        """)
        registros = cur.fetchall()
        for r in registros:
            grupo_nombre = r[4] or "Sin empresa"
            logs_por_grupo[grupo_nombre].append(r)
    except Exception:
        conn.rollback()
        app.logger.exception("Error cargando auditoria")
        flash_error(680)

    try:
        cur.execute("""
            SELECT
                m.id,
                m.fecha,
                COALESCE(u.usuario, '[eliminado]') AS usuario,
                e.nombre,
                m.item,
                m.accion,
                m.antes,
                m.despues,
                m.empresa
            FROM movimientos m
            LEFT JOIN usuarios u ON u.id = m.usuarioid
            LEFT JOIN empresas e ON e.id = m.empresa
            ORDER BY m.fecha DESC
            LIMIT 200
        """)
        movimientos = cur.fetchall()
        for m in movimientos:
            grupo_nombre = m[3] or "Sin empresa"
            movimientos_por_grupo[grupo_nombre].append(m)
    except Exception:
        conn.rollback()
        app.logger.exception("Error cargando movimientos")
        flash_error(681)

    try:
        cur.execute("""
            SELECT
                u.id,
                u.usuario,
                u.rol,
                ur.fecha,
                ur.empresa,
                ur.nombreempresa
            FROM usuarios u
            LEFT JOIN usuariosregistrados ur ON ur.usuarioid = u.id
            ORDER BY COALESCE(ur.fecha, SYSDATETIME()) DESC, ur.nombreempresa, u.usuario
        """)
        usuario_rows = cur.fetchall()
        agrupados = {}
        for user_id, usuario, rol, creado_en, grupo_id_item, grupo_nombre in usuario_rows:
            if user_id not in agrupados:
                agrupados[user_id] = {
                    "id": user_id,
                    "usuario": usuario,
                    "rol": rol,
                    "creado_en": creado_en or datetime.now(),
                    "grupos": [],
                    "grupo_ids": [],
                }
            if grupo_id_item is not None:
                agrupados[user_id]["grupos"].append(grupo_nombre)
                agrupados[user_id]["grupo_ids"].append(grupo_id_item)
        usuarios = [
            (
                data["id"],
                data["usuario"],
                data["rol"],
                data["creado_en"],
                ", ".join(data["grupos"]) if data["grupos"] else "Sin grupo",
                data["grupo_ids"],
            )
            for data in agrupados.values()
        ]

        cur.execute("""
            SELECT id, nombre
            FROM empresas
            WHERE archivado = 0 OR archivado IS NULL
            ORDER BY nombre
        """)
        grupos_todos = cur.fetchall()
    except Exception:
        conn.rollback()
        app.logger.exception("Error cargando usuarios/grupos")
        flash_error(682)

    cur.close()
    conn.close()

    return render_template(
        "admin_logs.html",
        registros=registros,
        movimientos=movimientos,
        logs_por_grupo=logs_por_grupo,
        movimientos_por_grupo=movimientos_por_grupo,
        usuarios=usuarios,
        grupos_todos=grupos_todos
    )


# ---------------- ADMIN: GRUPOS ----------------
@app.route("/admin/grupos", methods=["GET", "POST"])
def admin_grupos():
    return redirect(url_for("grupos"))


@app.route("/admin/grupos/abrir/<int:grupo_id>")
def admin_abrir_grupo(grupo_id):
    if not login_requerido():
        return redirect(url_for("login"))

    if not admin_requerido():
        return "Acceso denegado", 403

    session["grupo_id"] = grupo_id
    return redirect(url_for("archivo"))


# ---------------- ADMIN: MOVIMIENTOS ----------------
@app.route("/admin/movimientos")
def admin_movimientos():
    if not login_requerido():
        return redirect(url_for("login"))

    if not admin_requerido():
        return "Acceso denegado", 403

    return redirect(url_for("admin_logs"))


@app.route("/admin/movimientos/<int:mov_id>/deshacer", methods=["POST"])
def deshacer_movimiento(mov_id):
    if not login_requerido():
        return redirect(url_for("login"))

    if not admin_requerido():
        return "Acceso denegado", 403

    conn = get_db()
    cur = conn.cursor()

    cur.execute("""
        SELECT item, accion, antes, despues, empresa
        FROM movimientos
        WHERE id = %s
    """, (mov_id,))
    row = cur.fetchone()

    if not row:
        cur.close()
        conn.close()
        flash_error(683)
        return redirect(url_for("admin_movimientos"))

    item, accion, datos_antes, datos_despues, grupo_id = row
    entidad, entidad_id = parse_item_movimiento(item)
    datos_antes = json_or_none(datos_antes)
    datos_despues = json_or_none(datos_despues)

    if accion == "ELIMINAR_ARCHIVOS_CAJA":
        cur.close()
        conn.close()
        flash_error(684, fallback="Este movimiento no se puede deshacer.")
        return redirect(url_for("admin_movimientos"))

    try:
        if entidad == "archivo":
            if accion == "CREAR_ARCHIVO":
                cur.execute(
                    "DELETE FROM archivos WHERE id = %s AND grupo_id = %s",
                    (entidad_id, grupo_id)
                )
            elif accion == "ELIMINAR_ARCHIVO" and datos_antes:
                cur.execute("SELECT 1 FROM archivos WHERE id = %s", (datos_antes.get("id"),))
                if not cur.fetchone():
                    insert_with_identity(
                        cur,
                        "archivos",
                        """
                        INSERT INTO archivos (id, numero, nombre, caja_id, pdf_path, grupo_id, creado_por, tipo_doc)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                        """,
                        (
                            datos_antes.get("id"),
                            datos_antes.get("numero"),
                            datos_antes.get("nombre"),
                            datos_antes.get("caja_id"),
                            datos_antes.get("pdf_path"),
                            grupo_id,
                            None,
                            datos_antes.get("tipo_doc") or "CC",
                        )
                    )
            elif accion == "ARCHIVO_MOVER" and datos_antes:
                cur.execute(
                    "UPDATE archivos SET caja_id = %s WHERE id = %s AND grupo_id = %s",
                    (datos_antes.get("caja_id"), entidad_id, grupo_id)
                )
            elif accion == "MODIFICAR_ARCHIVO" and datos_antes:
                cur.execute(
                    """
                    UPDATE archivos
                    SET numero = %s, nombre = %s, caja_id = %s, pdf_path = %s
                    WHERE id = %s AND grupo_id = %s
                    """,
                    (
                        datos_antes.get("numero"),
                        datos_antes.get("nombre"),
                        datos_antes.get("caja_id"),
                        datos_antes.get("pdf_path"),
                        entidad_id,
                        grupo_id,
                    )
                )

        if entidad == "caja":
            if accion == "CREAR_CAJA":
                cur.execute(
                    "DELETE FROM cajas WHERE id = %s AND grupo_id = %s",
                    (entidad_id, grupo_id)
                )
            elif accion == "ELIMINAR_CAJA" and datos_antes:
                cur.execute("SELECT 1 FROM cajas WHERE id = %s", (entidad_id,))
                if not cur.fetchone():
                    insert_with_identity(
                        cur,
                        "cajas",
                        """
                        INSERT INTO cajas (id, rango_min, rango_max, grupo_id, is_pendiente)
                        VALUES (%s, %s, %s, %s, 0)
                        """,
                        (
                            entidad_id,
                            datos_antes.get("rango_min"),
                            datos_antes.get("rango_max"),
                            grupo_id,
                        )
                    )
            elif accion == "MODIFICAR_CAJA" and datos_antes:
                cur.execute(
                    """
                    UPDATE cajas
                    SET rango_min = %s, rango_max = %s
                    WHERE id = %s AND grupo_id = %s
                    """,
                    (
                        datos_antes.get("rango_min"),
                        datos_antes.get("rango_max"),
                        entidad_id,
                        grupo_id,
                    )
                )

        conn.commit()
        flash("Movimiento deshecho.", "success")
    except Exception as e:
        conn.rollback()
        flash_error(684, detail=str(e))
    finally:
        cur.close()
        conn.close()

    return redirect(url_for("admin_movimientos"))


# ---------------- ARCHIVADOR: TRANSFERIR ----------------
@app.route("/archivador/transferir", methods=["GET", "POST"])
def archivador_transferir():
    if not login_requerido():
        return redirect(url_for("login"))

    if request.method == "GET":
        return redirect(url_for("archivo") + "?view=especial")

    grupo_id = obtener_grupo_id()
    if not admin_requerido() or not es_archivador_grupo(grupo_id):
        return "Acceso denegado", 403

    cajas_ids_raw = request.form.get("cajas_ids", "")
    archivos_ids_raw = request.form.get("archivos_ids", "")
    grupo_destino = int(request.form.get("grupo_destino"))

    if es_archivador_grupo(grupo_destino):
        flash_error(208)
        return redirect(url_for("archivo") + "?view=especial")

    cajas_ids = [int(x) for x in cajas_ids_raw.split(",") if x.strip().isdigit()]
    archivos_ids = [int(x) for x in archivos_ids_raw.split(",") if x.strip().isdigit()]

    if not cajas_ids and not archivos_ids:
        flash_error(600)
        return redirect(url_for("archivo") + "?view=especial")

    conn = get_db()
    cur = conn.cursor()

    pendientes_id = asegurar_caja_sin_asignar(grupo_destino)
    warnings = []
    moved_file_ids = set()

    # Mover cajas completas
    for caja_id in cajas_ids:
        cur.execute(
            """
            SELECT rango_min, rango_max, is_pendiente
            FROM cajas
            WHERE id = %s AND grupo_id = %s
            """,
            (caja_id, grupo_id)
        )
        row = cur.fetchone()
        if not row:
            continue

        rmin, rmax, is_pendiente = row
        if is_pendiente:
            warnings.append(f"Caja {caja_id} es pendiente y no se movio.")
            continue

        # Aviso de solapamiento de rangos en destino
        cur.execute(
            """
            SELECT id, rango_min, rango_max
            FROM cajas
            WHERE grupo_id = %s
              AND is_pendiente = FALSE
              AND NOT (rango_max < %s OR rango_min > %s)
            """,
            (grupo_destino, rmin, rmax)
        )
        overlaps = cur.fetchall()
        if overlaps:
            warnings.append(
                f"Caja {caja_id} tiene solapamiento de rango en grupo destino."
            )

        cur.execute(
            """
            UPDATE cajas
            SET grupo_id = %s,
                grupo_origen_id = COALESCE(grupo_origen_id, %s)
            WHERE id = %s
            """,
            (grupo_destino, grupo_id, caja_id)
        )

        cur.execute(
            """
            UPDATE archivos
            SET grupo_id = %s,
                grupo_origen_id = COALESCE(grupo_origen_id, %s)
            WHERE caja_id = %s
            """,
            (grupo_destino, grupo_id, caja_id)
        )

        cur.execute("SELECT id FROM archivos WHERE caja_id = %s", (caja_id,))
        moved_file_ids.update([r[0] for r in cur.fetchall()])

    # Mover archivos sueltos
    for archivo_id in archivos_ids:
        if archivo_id in moved_file_ids:
            continue

        cur.execute(
            "SELECT numero FROM archivos WHERE id = %s AND grupo_id = %s",
            (archivo_id, grupo_id)
        )
        row = cur.fetchone()
        if not row:
            continue

        numero = row[0]
        cur.execute(
            """
            SELECT id
            FROM cajas
            WHERE grupo_id = %s
              AND is_pendiente = FALSE
              AND %s BETWEEN rango_min AND rango_max
            ORDER BY rango_min, id
            LIMIT 1
            """,
            (grupo_destino, numero)
        )
        dest = cur.fetchone()
        dest_caja = dest[0] if dest else pendientes_id

        cur.execute(
            """
            UPDATE archivos
            SET grupo_id = %s,
                caja_id = %s,
                grupo_origen_id = COALESCE(grupo_origen_id, %s)
            WHERE id = %s
            """,
            (grupo_destino, dest_caja, grupo_id, archivo_id)
        )

    conn.commit()
    cur.close()
    conn.close()

    if warnings:
        flash("Aviso: " + " | ".join(warnings), "info")
    else:
        flash("Elementos movidos correctamente.", "success")

    return redirect(url_for("archivo") + "?view=especial")


# ---------------- ARCHIVADOR: ELIMINAR ----------------
@app.route("/archivador/eliminar", methods=["GET", "POST"])
def archivador_eliminar():
    if not login_requerido():
        return redirect(url_for("login"))

    if request.method == "GET":
        return redirect(url_for("archivo") + "?view=especial")

    grupo_id = obtener_grupo_id()
    if not admin_requerido() or not es_archivador_grupo(grupo_id):
        return "Acceso denegado", 403

    cajas_ids_raw = request.form.get("cajas_ids", "")
    archivos_ids_raw = request.form.get("archivos_ids", "")

    cajas_ids = [int(x) for x in cajas_ids_raw.split(",") if x.strip().isdigit()]
    archivos_ids = [int(x) for x in archivos_ids_raw.split(",") if x.strip().isdigit()]

    if not cajas_ids and not archivos_ids:
        flash_error(600)
        return redirect(url_for("archivo") + "?view=especial")

    conn = get_db()
    cur = conn.cursor()

    # Eliminar archivos sueltos (que no esten en cajas seleccionadas)
    if archivos_ids:
        for chunk in iter_chunks(archivos_ids):
            in_clause, in_params = sqlserver_in_clause(chunk)
            cur.execute(
                f"DELETE FROM archivos WHERE id IN {in_clause} AND grupo_id = %s",
                (*in_params, grupo_id)
            )

    # Eliminar cajas y sus archivos
    for caja_id in cajas_ids:
        cur.execute(
            "DELETE FROM archivos WHERE caja_id = %s AND grupo_id = %s",
            (caja_id, grupo_id)
        )
        cur.execute(
            "DELETE FROM cajas WHERE id = %s AND grupo_id = %s",
            (caja_id, grupo_id)
        )

    conn.commit()
    cur.close()
    conn.close()

    flash("Elementos eliminados permanentemente.", "success")
    return redirect(url_for("archivo") + "?view=especial")

@app.route("/export/excel")
def export_excel():
    if not login_requerido():
        return redirect(url_for("login"))

    grupo_id = obtener_grupo_id()
    if not grupo_id:
        return redirect(url_for("grupos"))

    conn = get_db()
    cur = conn.cursor()

    # ===== 1) HOJA CAJAS (sin ID real, usando caja_visible) =====
    cur.execute("""
        WITH ranked AS (
            SELECT id, rango_min, rango_max,
                   ROW_NUMBER() OVER (ORDER BY rango_min, id) AS caja_visible
            FROM cajas
            WHERE grupo_id = %s AND is_pendiente = FALSE
        )
        SELECT
            r.caja_visible AS caja_num,
            r.rango_min,
            r.rango_max,
            COALESCE(COUNT(a.id),0) AS total_archivos
        FROM ranked r
        LEFT JOIN archivos a ON a.caja_id = r.id AND a.grupo_id = %s
        GROUP BY r.caja_visible, r.rango_min, r.rango_max
        ORDER BY r.caja_visible
    """, (grupo_id, grupo_id))
    cajas = cur.fetchall()

    # Caja pendiente (solo si tiene archivos)
    pendiente_id = asegurar_caja_sin_asignar(grupo_id)
    cur.execute(
        "SELECT COUNT(*) FROM archivos WHERE caja_id = %s AND grupo_id = %s",
        (pendiente_id, grupo_id)
    )
    total_caja0 = cur.fetchone()[0]

    # ===== 2) HOJA ARCHIVOS (caja_visible, documento, nombre, pdf si/no) =====
    cur.execute("""
        WITH ranked AS (
            SELECT id, ROW_NUMBER() OVER (ORDER BY rango_min, id) AS caja_visible
            FROM cajas
            WHERE grupo_id = %s AND is_pendiente = FALSE
        )
        SELECT
            CASE WHEN c.is_pendiente = 1 THEN 0 ELSE r.caja_visible END AS caja_num,
            a.numero,
            a.nombre,
            CASE WHEN a.pdf_path IS NULL OR a.pdf_path = '' THEN 'No' ELSE 'Si' END AS pdf
        FROM archivos a
        JOIN cajas c ON c.id = a.caja_id
        LEFT JOIN ranked r ON r.id = c.id
        WHERE a.grupo_id = %s
        ORDER BY caja_num, a.numero
    """, (grupo_id, grupo_id))
    archivos = cur.fetchall()

    cur.close()
    conn.close()

    # ===== Crear Excel =====
    wb = Workbook()

    # ---- Hoja 1: Cajas ----
    ws1 = wb.active
    ws1.title = "Cajas"
    ws1.append(["Caja", "Rango", "Total de archivos"])

    for (caja_num, rmin, rmax, total) in cajas:
        ws1.append([caja_num, f"{rmin}-{rmax}", total])

    if total_caja0 and total_caja0 > 0:
        # Caja 0 al final, sin rango
        ws1.append([0, "", total_caja0])

    # ---- Hoja 2: Archivos (en bloques por caja) ----
    ws2 = wb.create_sheet("Archivos")
    bloques_por_fila = 10
    encabezados_bloque = ["Caja", "Numero", "Nombre", "PDF"]
    separador = ""

    # Agrupar por caja
    archivos_por_caja = {}
    for (caja_num, numero, nombre, pdf) in archivos:
        archivos_por_caja.setdefault(caja_num, []).append((numero, nombre, pdf))

    cajas_orden = list(archivos_por_caja.keys())
    cajas_orden.sort(key=lambda x: (x == 0, x))

    for start in range(0, len(cajas_orden), bloques_por_fila):
        bloque_cajas = cajas_orden[start:start + bloques_por_fila]

        # Encabezados por bloque
        header_row = []
        for _ in bloque_cajas:
            header_row.extend(encabezados_bloque)
            header_row.append(separador)
        ws2.append(header_row)

        # Filas de datos (toma el maximo de cada bloque)
        max_len = 0
        for caja_num in bloque_cajas:
            max_len = max(max_len, len(archivos_por_caja.get(caja_num, [])))

        for idx in range(max_len):
            row = []
            for caja_num in bloque_cajas:
                items = archivos_por_caja.get(caja_num, [])
                if idx < len(items):
                    numero, nombre, pdf = items[idx]
                    row.extend([caja_num, numero, nombre, pdf])
                else:
                    row.extend(["", "", "", ""])
                row.append(separador)
            ws2.append(row)

        # Fila en blanco entre bloques
        ws2.append([])

    # Auto-width bÃ¡sico (opcional)
    for ws in [ws1, ws2]:
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 2, 40)

    # Guardar a memoria y enviar
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name="reporte_archivo.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


# PDF

@app.route("/pdf/<int:numero>")
def ver_pdf(numero):
    if not login_requerido():
        return redirect(url_for("login"))

    grupo_id = obtener_grupo_id()
    if not grupo_id:
        return redirect(url_for("grupos"))

    conn = get_db()
    cur = conn.cursor()
    cur.execute(
        "SELECT pdf_path FROM archivos WHERE numero = %s AND grupo_id = %s",
        (numero, grupo_id)
    )
    row = cur.fetchone()
    cur.close()
    conn.close()

    if not row or not row[0]:
        return error_text(420), 404

    filename = row[0]
    path = filename
    if not os.path.isabs(path):
        path = os.path.join(app.config["UPLOAD_FOLDER"], filename)
    if not os.path.exists(path):
        return error_text(421), 404

    return send_file(path, as_attachment=False, mimetype="application/pdf")


@app.route("/pdf/<int:numero>/pages")
def ver_pdf_paginas(numero):
    if not login_requerido():
        return redirect(url_for("login"))

    grupo_id = obtener_grupo_id()
    if not grupo_id:
        return redirect(url_for("grupos"))

    pages_raw = request.args.get("pages", "")
    download = request.args.get("download") == "1"
    pages = []
    for p in pages_raw.split(","):
        p = p.strip()
        if p.isdigit():
            pages.append(int(p) - 1)

    if not pages:
        return error_text(422, fallback="No hay paginas seleccionadas."), 400

    if PdfReader is None or PdfWriter is None:
        return error_text(904, fallback="No se pudo procesar el PDF seleccionado."), 500

    conn = get_db()
    cur = conn.cursor()
    cur.execute(
        "SELECT pdf_path FROM archivos WHERE numero = %s AND grupo_id = %s",
        (numero, grupo_id)
    )
    row = cur.fetchone()
    cur.close()
    conn.close()

    if not row or not row[0]:
        return error_text(420), 404

    filename = row[0]
    path = os.path.join(app.config["UPLOAD_FOLDER"], filename)
    if not os.path.exists(path):
        return error_text(421), 404

    try:
        reader = PdfReader(path)
        writer = PdfWriter()
        total = len(reader.pages)
        for idx in pages:
            if 0 <= idx < total:
                writer.add_page(reader.pages[idx])
    except Exception:
        app.logger.exception("Error creando PDF seleccionado")
        return error_text(904, fallback="No se pudo procesar el PDF seleccionado."), 500

    output = BytesIO()
    writer.write(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=download,
        download_name=f"doc_{numero}_seleccion.pdf",
        mimetype="application/pdf"
    )

if __name__ == "__main__":
    app.run(debug=False)

@app.errorhandler(404)
def handle_not_found(e):
    if login_requerido():
        flash_error(750)
        return redirect(request.referrer or url_for("inicio"))
    return error_text(750), 404


@app.errorhandler(405)
def handle_method_not_allowed(e):
    if login_requerido():
        flash_error(751)
        return redirect(request.referrer or url_for("inicio"))
    return error_text(751), 405


@app.errorhandler(413)
def handle_request_too_large(e):
    mensaje = "La carga supera el limite permitido del servidor. Para PDF masivo, el maximo es 500 archivos y 250 MB por lote."
    if login_requerido():
        flash(mensaje, "error")
        return redirect(request.referrer or url_for("archivo"))
    return mensaje, 413


@app.errorhandler(Exception)
def handle_unhandled_exception(e):
    if isinstance(e, HTTPException):
        return e
    app.logger.exception("Unhandled exception")
    if login_requerido():
        flash_error(900)
        return redirect(request.referrer or url_for("inicio"))
    return error_text(900), 500
